import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

print("=== TOTAL REGISTROS EN EL EXCEL ORIGINAL ===")
total_unique_names = set()

for sheet in wb.sheetnames:
    ws = wb[sheet]
    count = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and any(r):
            count += 1
            name_candidate = str(r[1] if len(r) > 1 and r[1] else r[0] if len(r) > 0 and r[0] else '').strip()
            if name_candidate and len(name_candidate) > 2 and name_candidate.lower() not in ('fullname', 'nombre', 'person a', 'person b'):
                total_unique_names.add(name_candidate.lower())
    print(f"• Pestaña '{sheet}': {count} registros")

print(f"\nTOTAL NOMBRES DE PERSONAS ÚNICAS EN TODO EL EXCEL: {len(total_unique_names)}")
