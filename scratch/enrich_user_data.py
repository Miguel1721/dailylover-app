import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

matches_sheets = [s for s in wb.sheetnames if 'MATCHES' in s and s not in ('MISSING MATCHES', 'TROUBLE MATCHES')]

person_cities = {}
person_responsables = {}
person_plans = {}

for sname in matches_sheets:
    ws = wb[sname]
    row1 = [str(c).upper() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    city_idx = -1
    plan_idx = -1
    pa_idx = -1
    pb_idx = -1
    
    for idx, h in enumerate(row1):
        if 'CITY' in h or 'CIUDAD' in h: city_idx = idx
        if 'PLAN' in h: plan_idx = idx
        if 'PERSON A' in h: pa_idx = idx
        if 'PERSON B' in h: pb_idx = idx
    
    # Psychologist name from sheet name
    psych = sname.replace('MATCHES', '').strip()
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        city_val = str(row[city_idx]).strip() if city_idx >= 0 and city_idx < len(row) and row[city_idx] else ""
        plan_val = str(row[plan_idx]).strip() if plan_idx >= 0 and plan_idx < len(row) and row[plan_idx] else ""
        
        pa_val = str(row[pa_idx]).strip() if pa_idx >= 0 and pa_idx < len(row) and row[pa_idx] else ""
        pb_val = str(row[pb_idx]).strip() if pb_idx >= 0 and pb_idx < len(row) and row[pb_idx] else ""
        
        for p in (pa_val, pb_val):
            if not p or len(p) < 3 or p.lower() in ('person a', 'person b', 'none', 'null'): continue
            p_key = p.lower()
            
            if psych:
                person_responsables[p_key] = psych
                
            if city_val and city_val.lower() not in ('none', 'null', 'city', 'ciudad'):
                c_clean = city_val.lower()
                if 'bog' in c_clean or 'bgta' in c_clean:
                    person_cities[p_key] = 'Bogotá'
                elif 'med' in c_clean or 'mde' in c_clean:
                    person_cities[p_key] = 'Medellín'
                elif 'cali' in c_clean:
                    person_cities[p_key] = 'Cali'
                elif 'barranq' in c_clean:
                    person_cities[p_key] = 'Barranquilla'
                else:
                    person_cities[p_key] = city_val

print(f"EXTRACTED CITIES FOR {len(person_cities)} INDIVIDUALS FROM PSYCHOLOGIST MATCHES SHEETS!")
print(f"EXTRACTED RESPONSABLES FOR {len(person_responsables)} INDIVIDUALS FROM PSYCHOLOGIST MATCHES SHEETS!")
