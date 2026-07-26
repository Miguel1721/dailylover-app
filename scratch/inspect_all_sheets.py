import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

print("=== INSPECTING ALL SHEETS FOR CITIES AND RESPONSABLES ===")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [str(h) for h in row1 if h is not None]
    print(f"\nSheet '{sheet}' ({ws.max_row} rows):")
    print("  Headers:", headers[:8])
