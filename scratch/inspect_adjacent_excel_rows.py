import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', data_only=True)

targets = ['alvaro jose echeverry', 'oscar rojas', 'alexander restrepo', 'gerard jesus suarez']

print("=== REVISANDO CELDAS ADYACENTES (ARRIBA, ABAJO, DERECHA) EN PESTAÑAS MATCHES ===")
for sheet in wb.sheetnames:
    if 'MATCHES' not in sheet: continue
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    for r_idx in range(len(rows)):
        r = rows[r_idx]
        if not r: continue
        line = ' '.join([str(c) for c in r if c is not None]).lower()
        if any(t in line for t in targets):
            above = ' '.join([str(c) for c in rows[r_idx-1] if c is not None]) if r_idx > 0 else 'N/A'
            below = ' '.join([str(c) for c in rows[r_idx+1] if c is not None]) if r_idx < len(rows)-1 else 'N/A'
            print(f"Sheet '{sheet}' Fila {r_idx+1}:")
            print(f"  • Fila Arriba ({r_idx}): {above[:120]}")
            print(f"  • Fila Actual ({r_idx+1}): {line[:120]}")
            print(f"  • Fila Abajo  ({r_idx+2}): {below[:120]}\n")
