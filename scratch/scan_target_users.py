import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

target_names = ['ruben rojas', 'carolina dorado', 'manuel blanco']

print("=== SCANNING ALL SHEETS FOR TARGET USERS ===")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r_idx, r in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        line = ' '.join([str(c) for c in r if c is not None]).lower()
        if any(t in line for t in target_names):
            print(f"Sheet '{sheet}' (row {r_idx}): {line[:120]}")
