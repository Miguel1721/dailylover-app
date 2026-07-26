import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

target_names = ['ruben rojas', 'carolina dorado', 'manuel blanco']

print("=== CHECKING PROFILES SHEET ===")
ws_prof = wb['PROFILES']
headers = [str(c) for c in next(ws_prof.iter_rows(min_row=1, max_row=1, values_only=True))]
print("PROFILES Col 8:", headers[8] if len(headers) > 8 else "N/A")

for r in ws_prof.iter_rows(min_row=2, values_only=True):
    name = str(r[1] or '').strip().lower()
    if any(t in name for t in target_names):
        age_val = r[8] if len(r) > 8 else None
        resp_val = r[3] if len(r) > 3 else None
        city_val = r[9] if len(r) > 9 else (r[4] if len(r) > 4 else None)
        print(f"PROFILES -> Name: '{r[1]}' | Age (Col 8): '{age_val}' | Resp (Col 3): '{resp_val}' | City: '{city_val}'")

print("\n=== CHECKING CLIENTS PLANS SHEET ===")
ws_plans = wb['Clients plans']
plans_config = [('VIP 195k', 0, 1), ('Premium 150k', 2, 3), ('Estándar Plus 98k', 4, 5), ('Estándar 65k (2 citas)', 6, 7), ('Estándar 65k (1 cita)', 8, 9), ('Básico 40k', 10, 11)]

found_plans = 0
for r in ws_plans.iter_rows(min_row=2, values_only=True):
    if not r: continue
    for p_name, n_idx, e_idx in plans_config:
        n_val = str(r[n_idx]).strip().lower() if n_idx < len(r) and r[n_idx] else ''
        if any(t in n_val for t in target_names):
            found_plans += 1
            print(f"CLIENTS PLANS -> Name: '{r[n_idx]}' | Plan: '{p_name}'")

if found_plans == 0:
    print("Ninguno de estos usuarios aparece en la pestaña 'Clients plans'")
