import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', data_only=True)

valid_dates = []

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r_idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not r: continue
        for c in r:
            if hasattr(c, 'year'):
                # Ignore Excel US/UK inverted dates where month/day were swapped (month > 7 in 2026)
                if c.year == 2026 and c.month <= 7:
                    valid_dates.append((sheet, r_idx, c))

valid_dates.sort(key=lambda x: x[2])

print(f"TOTAL FECHAS VALIDAS: {len(valid_dates)}")
if valid_dates:
    latest = valid_dates[-1]
    print(f"ULTIMA FECHA REAL REGISTRADA EN EL EXCEL: {latest[2].strftime('%d/%m/%Y')} (Pestaña: '{latest[0]}', Fila {latest[1]})")

    print("\nULTIMAS 5 FECHAS REGISTRADAS EN EL EXCEL:")
    for sheet, row, d in valid_dates[-5:]:
        print(f"  • {d.strftime('%d/%m/%Y')} en pestaña '{sheet}' (Fila {row})")
