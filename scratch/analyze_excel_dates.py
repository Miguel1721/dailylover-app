import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', data_only=True)

dates = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        for cell in r:
            if cell is not None and hasattr(cell, 'year'):
                if 2018 <= cell.year <= 2026:
                    dates.append((sheet, cell))

dates.sort(key=lambda x: x[1])

print(f"TOTAL DATES FOUND IN EXCEL: {len(dates)}")
if dates:
    print(f"EARLIEST DATE: {dates[0][1].strftime('%d/%m/%Y')} (Sheet: '{dates[0][0]}')")
    print(f"LATEST DATE: {dates[-1][1].strftime('%d/%m/%Y')} (Sheet: '{dates[-1][0]}')")

    # Year breakdown
    years = {}
    for sheet, d in dates:
        y = d.year
        years[y] = years.get(y, 0) + 1
    print("\nDISTRIBUCION DE REGISTROS POR AÑO EN EL EXCEL:")
    for y in sorted(years.keys()):
        print(f"  • Año {y}: {years[y]} fechas de registro/citas")
