import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

targets = ['daniela ballesteros', 'alexis palacio', 'camilo humberto prieto', 'diego alejandro neira', 'eduardo espinosa']

print("=== CHECKING TARGET USERS ACROSS ALL SHEETS IN EXCEL ===")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r_idx, r in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not r: continue
        line = ' '.join([str(c) for c in r if c is not None])
        line_low = line.lower()
        if any(t in line_low for t in targets):
            print(f"Sheet '{sheet}' (Row {r_idx}): {line[:140]}")
