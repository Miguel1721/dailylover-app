import openpyxl
import re

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', data_only=True)

names_in_sheet = {}

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r: continue
        for c in r:
            if c and isinstance(c, str):
                s = c.strip()
                # Check if looks like a name (2-4 words, capitalized, no notes keywords)
                if 5 < len(s) < 40 and not any(k in s.lower() for k in ['waitlist', 'espera', 'creo', 'salió', 'rarito', 'date', 'pago', 'refund', 'trouble']):
                    words = s.split()
                    if 2 <= len(words) <= 5 and all(w[0].isupper() for w in words if w[0].isalpha()):
                        names_in_sheet[s.lower()] = s

print(f"COLLECTED {len(names_in_sheet)} KNOWN CLEAN NAMES FROM EXCEL.")
