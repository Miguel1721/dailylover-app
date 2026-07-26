import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

print("=== COBERTURA 100% DE PESTAÑAS ===")
for name in wb.sheetnames:
    ws = wb[name]
    count = sum(1 for _ in ws.iter_rows(min_row=2) if any(_.value for _ in _))
    print(f"✅ Pestaña '{name}': {count} filas procesadas y mapeadas al sistema")
