import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', data_only=True)

targets = ['santiago botero', 'maria paula forero', 'gabriela supelano']

print("=== CHECKING SANTIAGO BOTERO GIRALDO IN MATCHES TABS ===")
for sheet in wb.sheetnames:
    if 'MATCHES' not in sheet: continue
    ws = wb[sheet]
    for r_idx, r in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not r: continue
        line = ' '.join([str(c) for c in r if c is not None]).lower()
        if any(t in line for t in targets):
            clean_line = line.encode('ascii', 'ignore').decode('ascii')
            print(f"Sheet '{sheet}' (Row {r_idx}): {clean_line[:140]}")
