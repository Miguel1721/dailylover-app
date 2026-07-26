import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

matches_sheets = [s for s in wb.sheetnames if 'MATCHES' in s and s not in ('MISSING MATCHES', 'TROUBLE MATCHES')]

person_cities = {}
person_plans = {}

for sname in matches_sheets:
    ws = wb[sname]
    row1 = [str(c).upper() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    
    city_idx = -1
    plan_idx = -1
    pa_idx = -1
    pb_idx = -1
    
    for idx, h in enumerate(row1):
        if 'CITY' in h or 'CIUDAD' in h: city_idx = idx
        if 'PLAN' in h: plan_idx = idx
        if 'PERSON A' in h: pa_idx = idx
        if 'PERSON B' in h: pb_idx = idx
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        city_val = str(row[city_idx]).strip() if city_idx >= 0 and city_idx < len(row) and row[city_idx] else ""
        plan_val = str(row[plan_idx]).strip() if plan_idx >= 0 and plan_idx < len(row) and row[plan_idx] else ""
        
        pa_val = str(row[pa_idx]).strip() if pa_idx >= 0 and pa_idx < len(row) and row[pa_idx] else ""
        pb_val = str(row[pb_idx]).strip() if pb_idx >= 0 and pb_idx < len(row) and row[pb_idx] else ""
        
        if city_val and city_val.lower() not in ('none', 'null', 'city', 'ciudad'):
            if pa_val: person_cities[pa_val.lower()] = city_val
            if pb_val: person_cities[pb_val.lower()] = city_val
            
        if plan_val and plan_val.lower() not in ('none', 'null', 'plan'):
            if pa_val: person_plans[pa_val.lower()] = plan_val
            if pb_val: person_plans[pb_val.lower()] = plan_val

print(f"EXTRA CITIES EXTRACTED FROM MATCHES SHEETS: {len(person_cities)} unique people with city")
print(f"EXTRA PLANS EXTRACTED FROM MATCHES SHEETS: {len(person_plans)} unique people with plan")
print("Sample extracted cities:", list(person_cities.items())[:10])
print("Sample extracted plans:", list(person_plans.items())[:10])
