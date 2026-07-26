import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

print("=== COBERTURA 100% DE PESTANAS EN EL SISTEMA ===")
total_filas = 0
for name in wb.sheetnames:
    ws = wb[name]
    count = sum(1 for row in ws.iter_rows(min_row=2) if any(c.value for c in row))
    total_filas += count
    print(f"-> Pestana '{name}': {count} filas mapeadas")

print(f"\nTOTAL FILAS PROCESADAS DE TODO EL EXCEL: {total_filas}")
