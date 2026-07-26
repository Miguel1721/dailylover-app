import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', data_only=True)

targets = ['alvaro jose echeverry', 'oscar rojas', 'mariana botero', 'daniel ospina', 'nestor javier almanza', 'juan ricci', 'manolo montoya', 'diana', 'juliana', 'andrea', 'laura', 'catalina', 'camila', 'maria', 'paula', 'sofia', 'valentina', 'carolina']

print("=== CHECKING MATCHES TABS FOR SPECIFIC ROWS ===")
for sheet in wb.sheetnames:
    if 'MATCHES' not in sheet: continue
    ws = wb[sheet]
    for r_idx, r in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        if not r: continue
        line = ' '.join([str(c) for c in r if c is not None]).lower()
        if 'echeverry' in line or 'oscar rojas' in line or 'mariana botero' in line or 'daniel ospina' in line or 'waitlist' in line:
            clean_line = line.encode('ascii', 'ignore').decode('ascii')
            print(f"Sheet '{sheet}' Row {r_idx}: {clean_line[:140]}")
