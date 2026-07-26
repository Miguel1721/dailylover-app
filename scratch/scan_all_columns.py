import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

print("=== ALL COLUMNS IN ALL SHEETS ===")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    try:
        row1 = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        headers = [f"Col {i}: {h}" for i, h in enumerate(row1) if h is not None]
        print(f"\n--- Sheet '{sheet}' ---")
        for h in headers:
            print("  ", h)
    except StopIteration:
        pass
