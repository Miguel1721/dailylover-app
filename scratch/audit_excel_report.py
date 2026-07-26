import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\jeloz\Downloads\Daily Lover MATCHMAKING.xlsx', read_only=True)

print("=== AUDITORIA COMPLETA DE PESTAÑAS Y REGISTROS DEL EXCEL ===")
total_rows_excel = 0
sheet_summary = []
for sheet in wb.sheetnames:
    ws = wb[sheet]
    count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and any(r))
    total_rows_excel += count
    sheet_summary.append((sheet, count))
    print(f"• Pestaña '{sheet}': {count} filas con datos")

print(f"\nTOTAL FILAS EN TODO EL EXCEL: {total_rows_excel}")
