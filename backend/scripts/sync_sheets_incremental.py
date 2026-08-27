"""
Incremental Sync from Google Sheets to Daily Lover Database (Pull Worker)
Runs incrementally every 15 minutes or on demand.
Repointable via GOOGLE_SHEET_MIGRATION_ID environment variable.
"""

import os
import io
import re
import urllib.request
import logging
import asyncio
from datetime import datetime
from typing import Dict, Any, List, Optional
import openpyxl
from sqlalchemy import text
from app.database import AsyncSessionLocal
from app.config import get_settings

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("sync_sheets_incremental")

settings = get_settings()
SHEET_ID = os.environ.get("GOOGLE_SHEET_MIGRATION_ID") or getattr(settings, "google_sheet_migration_id", "1ziZsPwYv6I3fEIEyVM7I0Na7LLgzrwyfM8vcglUQ8RA")
SHEET_EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

ACTIVE_PSYCHOLOGISTS = [
    "JENN", "ANA", "SILVI", "STEFFY", "SOFI", "MAPE D", "ALEJA", "MANU 1", "MANU 2", "PIA"
]

def extract_crm_id_from_cell(cell) -> Optional[str]:
    """Extrae el ID del CRM desde el hipervínculo de la celda o fórmula."""
    target = None
    if cell.hyperlink and cell.hyperlink.target:
        target = cell.hyperlink.target
    elif cell.value and isinstance(cell.value, str) and "=HYPERLINK(" in cell.value.upper():
        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value, re.IGNORECASE)
        if m:
            target = m.group(1)
    
    if target:
        m_id = re.search(r"(?:client|profile|view)[/=](\d+)", target, re.IGNORECASE)
        if m_id:
            return m_id.group(1)
        m_q = re.search(r"[?&]id=(\d+)", target, re.IGNORECASE)
        if m_q:
            return m_q.group(1)
    return None

def is_valid_person_name(val: Any) -> bool:
    if not val:
        return False
    s = str(val).strip()
    if not s or len(s) < 3:
        return False
    if re.match(r'^\d{4}-\d{2}-\d{2}', s) or re.match(r'^\d{1,2}/\d{1,2}/\d{4}', s):
        return False
    s_upper = s.upper()
    invalid_keywords = [
        "PIDIO REFOUND", "REFUND", "SLOT", "SLOTS", "HISTÓRICO", "HISTORICO",
        "NAME", "NOMBRE", "PERSONA A", "PERSONA B", "STATUS", "NO TIENE",
        "DESCALIFICADO", "TROUBLE", "SIN GENTE", "TRUE", "FALSE", "VERDADERO"
    ]
    if any(k in s_upper for k in invalid_keywords):
        return False
    if not re.search(r'[a-zA-ZáéíóúÁÉÍÓÚñÑ]', s):
        return False
    return True

def normalize_city(raw_city: Optional[str]) -> str:
    if not raw_city:
        return ""
    c = str(raw_city).strip()
    if not c:
        return ""
    c_lower = c.lower()
    if any(noise in c_lower for noise in [
        "slot", "exist", "hist", "jenn", "silvi", "ana", "steffy", "sofi", "aleja", "pia", "manu", "mape",
        "none", "null", "nan", ",", "2 dates", "todo el mundo", "refound", "refund", "true", "false", "status"
    ]):
        return ""
    mapping = {
        "bgota": "Bogotá", "bogota": "Bogotá", "bog": "Bogotá",
        "medellin": "Medellín", "med": "Medellín",
        "baq": "Barranquilla", "bquilla": "Barranquilla", "quilla": "Barranquilla",
        "bmanga": "Bucaramanga", "buc": "Bucaramanga", "buca": "Bucaramanga",
        "ctg": "Cartagena", "cartagena": "Cartagena",
        "mad": "Madrid", "madrid": "Madrid",
        "mia": "Miami", "miami": "Miami",
        "cdmx": "CDMX", "mexico": "CDMX", "méxico": "CDMX",
        "peira": "Pereira", "pereira": "Pereira",
        "ibag": "Ibagué", "ibague": "Ibagué", "cali": "Cali", "caqu": "Caquetá"
    }
    for k, v in mapping.items():
        if k in c_lower:
            return v
    return c.title()

def normalize_pref(val: Optional[str]) -> str:
    if not val:
        return "hetero"
    v = val.lower().strip()
    if "bi" in v:
        return "bi"
    if "lesb" in v:
        return "lesb"
    if "gay" in v or "homo" in v:
        return "gay"
    return "hetero"

def normalize_plan(raw_plan: Optional[str]) -> str:
    if not raw_plan:
        return ""
    p = str(raw_plan).lower().strip()
    if "vip" in p or "195k" in p or "295k" in p:
        return "VIP 195k"
    if "2 date" in p or "2 cita" in p or "standard" in p or "estandar" in p or "estándar" in p or "65k" in p or "98k" in p or "150k" in p or "premium" in p:
        return "Estándar 65k (2 citas)"
    if "1 date" in p or "1 cita" in p or "basic" in p or "basico" in p or "básico" in p or "40k" in p:
        return "Básico 40k"
    return ""

def get_slots_count(plan_name: str) -> int:
    if plan_name == "VIP 195k":
        return 4
    elif plan_name == "Estándar 65k (2 citas)":
        return 3
    elif plan_name == "Básico 40k":
        return 2
    return 2  # Default reserva cuando está pendiente plan

async def sync_incremental():
    logger.info(f"Iniciando sincronización incremental desde Google Sheet ID: {SHEET_ID}")
    
    # 1. Descargar XLSX
    req = urllib.request.Request(SHEET_EXPORT_URL, headers={"User-Agent": "Mozilla/5.0 DailyLoverSync/1.0"})
    try:
        with urllib.request.urlopen(req, timeout=45) as resp:
            content = resp.read()
        logger.info(f"XLSX descargado exitosamente ({len(content)} bytes)")
    except Exception as e:
        logger.error(f"Error al descargar Google Sheet: {e}")
        return

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    
    # Mapa de Planes y Preferencias desde la pestaña 'Clients plans' si existe en el Sheet
    clients_plans_map = {}
    if "Clients plans" in wb.sheetnames:
        ws_cp = wb["Clients plans"]
        for r in range(2, ws_cp.max_row + 1):
            cp_name = str(ws_cp.cell(r, 2).value or "").strip().lower()
            cp_plan = str(ws_cp.cell(r, 5).value or "").strip()
            cp_city = str(ws_cp.cell(r, 6).value or "").strip()
            cp_pref = str(ws_cp.cell(r, 7).value or "").strip()
            if cp_name:
                clients_plans_map[cp_name] = {
                    "plan": normalize_plan(cp_plan),
                    "city": normalize_city(cp_city),
                    "pref": normalize_pref(cp_pref)
                }

    async with AsyncSessionLocal() as db:
        new_clients_count = 0
        new_matches_count = 0
        updated_crm_ids = 0

        # 2. Ingesta de Perfiles desde PROFILES (Estructura Copia final: No. | FullName | FECHA | Responsable | Ciudad y años | SLOTS CREADOS)
        if "PROFILES" in wb.sheetnames:
            ws = wb["PROFILES"]
            logger.info(f"Procesando pestaña PROFILES ({ws.max_row} filas)...")
            
            # Detectar encabezados dinámicamente con fallback exacto a Copia Final
            headers_map = {}
            for c in range(1, ws.max_column + 1):
                h_val = str(ws.cell(row=1, column=c).value or "").strip().upper()
                if h_val:
                    headers_map[h_val] = c
            
            col_name = headers_map.get("FULLNAME") or headers_map.get("NOMBRE") or 2
            col_fecha = headers_map.get("FECHA") or 3
            col_psyc = headers_map.get("RESPONSABLE") or headers_map.get("PSICOLOGA") or 4
            col_city = headers_map.get("CIUDAD Y AÑOS") or headers_map.get("CIUDAD") or 5
            col_slots = headers_map.get("SLOTS CREADOS") or headers_map.get("SLOTS") or 6

            for row_idx in range(2, ws.max_row + 1):
                cell_name = ws.cell(row=row_idx, column=col_name)
                raw_name = str(cell_name.value or "").strip()
                
                # Validación estricta de nombre de persona
                if not is_valid_person_name(raw_name):
                    continue
                
                crm_id = extract_crm_id_from_cell(cell_name)
                raw_psyc = str(ws.cell(row=row_idx, column=col_psyc).value or "").strip().upper()
                raw_city = str(ws.cell(row=row_idx, column=col_city).value or "").strip()
                
                # REGLA: Si la psicóloga no es una de las 10 oficiales, NO asignar a SILVI silenciosamente.
                # Dejar vacío "" para revisión manual.
                psyc = raw_psyc if raw_psyc in ACTIVE_PSYCHOLOGISTS else ""
                
                # Buscar Plan y Preferencia en mapa de 'Clients plans' si existe; de lo contrario dejar vacío ""
                cp_data = clients_plans_map.get(raw_name.lower(), {})
                city = (normalize_city(raw_city) or cp_data.get("city") or "")[:50]
                pref = cp_data.get("pref") or ""
                plan = cp_data.get("plan") or ""

                # Verificar si el cliente ya existe en users/profiles
                res = await db.execute(text("""
                    SELECT u.id, u.crm_id, p.responsable, p.plan_tier, p.orientation, p.city 
                    FROM users u
                    LEFT JOIN profiles p ON p.user_id = u.id
                    WHERE LOWER(TRIM(u.name)) = LOWER(TRIM(:name))
                    LIMIT 1
                """), {"name": raw_name})
                user_row = res.fetchone()

                if user_row:
                    uid, existing_crm = user_row.id, user_row.crm_id
                    if crm_id and existing_crm != crm_id:
                        await db.execute(text("UPDATE users SET crm_id = :cid WHERE id = :uid"), {"cid": crm_id, "uid": uid})
                        updated_crm_ids += 1
                    # Usar datos existentes de perfil si en el sheet estaban vacíos
                    if not plan and user_row.plan_tier:
                        plan = user_row.plan_tier
                    if not pref and user_row.orientation:
                        pref = user_row.orientation
                    if not city and user_row.city:
                        city = user_row.city
                    if not psyc and user_row.responsable:
                        psyc = user_row.responsable
                else:
                    # Crear usuario y perfil nuevo
                    placeholder_phone = f"GEN_{row_idx}_{int(datetime.utcnow().timestamp())}"
                    res_ins = await db.execute(text("""
                        INSERT INTO users (name, phone, crm_id, created_at)
                        VALUES (:name, :phone, :cid, NOW())
                        RETURNING id
                    """), {"name": raw_name, "phone": placeholder_phone, "cid": crm_id})
                    uid = res_ins.scalar()
                    await db.execute(text("""
                        INSERT INTO profiles (user_id, city, orientation, plan_tier, responsable, updated_at)
                        VALUES (:uid, :city, :pref, :plan, :resp, NOW())
                        ON CONFLICT (user_id) DO NOTHING
                    """), {"uid": uid, "city": city, "pref": pref, "plan": plan, "resp": psyc})
                    new_clients_count += 1

                # Verificar si ya tiene slots creados en operational_matches
                res_slots = await db.execute(text("""
                    SELECT COUNT(*) FROM operational_matches
                    WHERE LOWER(TRIM(person_a)) = LOWER(TRIM(:name))
                """), {"name": raw_name})
                if res_slots.scalar() == 0:
                    num_slots = get_slots_count(plan)
                    initial_status = "PENDIENTE" if plan else "PENDIENTE PLAN"
                    for s in range(1, num_slots + 1):
                        await db.execute(text("""
                            INSERT INTO operational_matches 
                            (city, pref, plan_tier, person_a, psychologist_name, slot_number, status, person_a_crm_id, created_at, updated_at)
                            VALUES (:city, :pref, :plan, :pA, :psyc, :slot, :st, :cid, NOW(), NOW())
                        """), {
                            "city": city, "pref": pref, "plan": plan, "pA": raw_name,
                            "psyc": psyc, "slot": s, "st": initial_status, "cid": crm_id
                        })
                    new_matches_count += num_slots

            await db.commit()

        logger.info(f"Sincronización incremental finalizada con éxito.")
        logger.info(f"Resumen: Nuevos Clientes: {new_clients_count}, Nuevos Slots: {new_matches_count}, CRM IDs Actualizados: {updated_crm_ids}")

if __name__ == "__main__":
    asyncio.run(sync_incremental())
