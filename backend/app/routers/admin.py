from fastapi import APIRouter, Depends, Query, HTTPException, Header, Request
from pydantic import BaseModel
from datetime import datetime, timedelta
import os

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from app.database import get_db
from app.config import get_settings
from app.core.permissions import require_permission
from typing import Optional
import math
import json

router = APIRouter(prefix="/api/v1/admin", tags=["Admin"])


import re

MONTH_NAMES_ES = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Abr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dic"
}


def clean_excel_date_str(raw):
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("null", "none", "nan", "por agendar"):
        return None

    # 1. Si es serial numérico de Excel (ej: 45400 o 45400.0)
    try:
        val = float(s)
        if 30000 <= val <= 70000:
            dt = datetime(1899, 12, 30) + timedelta(days=val)
            return dt.strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        pass

    # 2. Si es un formato de texto tipo "4.25 7pm", "4.25", "4/25 7:00pm", "25/4 7pm"
    m = re.search(r'(\d{1,2})[\./-](\d{1,2})(?:\s+(.*))?', s, re.IGNORECASE)
    if m:
        p1 = int(m.group(1))
        p2 = int(m.group(2))
        time_part = (m.group(3) or "").strip()

        # Determinar cuál es mes y cuál es día
        if p1 <= 12 and p2 <= 31:
            month_num, day_num = p1, p2
        elif p2 <= 12 and p1 <= 31:
            month_num, day_num = p2, p1
        else:
            return s

        if 1 <= month_num <= 12 and 1 <= day_num <= 31:
            month_name = MONTH_NAMES_ES.get(month_num, f"{month_num:02d}")
            formatted = f"{day_num} {month_name}"
            if time_part:
                clean_time = time_part.upper().replace('.', '')
                formatted += f", {clean_time}"
            return formatted

    return s



# ─── DATA DIAGNOSTICS ─────────────────────────────────────────────────────────

@router.get("/data-health")
async def get_data_health(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("dashboard", "view"))
):
    """Diagnóstico de salud de datos — cuántos matches están vinculados por ID vs nombre."""
    total_matches = (await db.execute(text("SELECT COUNT(*) FROM historical_matches"))).scalar() or 0
    linked_a = (await db.execute(text("SELECT COUNT(*) FROM historical_matches WHERE user_id_a IS NOT NULL"))).scalar() or 0
    linked_b = (await db.execute(text("SELECT COUNT(*) FROM historical_matches WHERE user_id_b IS NOT NULL"))).scalar() or 0
    both_linked = (await db.execute(text("SELECT COUNT(*) FROM historical_matches WHERE user_id_a IS NOT NULL AND user_id_b IS NOT NULL"))).scalar() or 0
    orphan_a = (await db.execute(text("SELECT COUNT(*) FROM historical_matches WHERE user_id_a IS NULL AND person_a IS NOT NULL"))).scalar() or 0
    orphan_b = (await db.execute(text("SELECT COUNT(*) FROM historical_matches WHERE user_id_b IS NULL AND person_b IS NOT NULL"))).scalar() or 0

    # Personas en matches que NO existen en users
    unregistered_res = await db.execute(text("""
        SELECT DISTINCT name FROM (
            SELECT hm.person_a AS name FROM historical_matches hm
            WHERE hm.user_id_a IS NULL AND hm.person_a IS NOT NULL
            AND NOT EXISTS (SELECT 1 FROM users u WHERE unaccent(lower(trim(u.name))) = unaccent(lower(trim(hm.person_a))))
            UNION
            SELECT hm.person_b AS name FROM historical_matches hm
            WHERE hm.user_id_b IS NULL AND hm.person_b IS NOT NULL
            AND NOT EXISTS (SELECT 1 FROM users u WHERE unaccent(lower(trim(u.name))) = unaccent(lower(trim(hm.person_b))))
        ) orphans
        ORDER BY name
        LIMIT 30
    """))
    unregistered_names = [r.name for r in unregistered_res.fetchall()]

    # Usuarios con client_code
    total_users = (await db.execute(text("SELECT COUNT(*) FROM users"))).scalar() or 0
    with_code = (await db.execute(text("SELECT COUNT(*) FROM users WHERE client_code IS NOT NULL"))).scalar() or 0
    with_cedula = (await db.execute(text("SELECT COUNT(*) FROM users WHERE id_number IS NOT NULL"))).scalar() or 0

    # Status distribution
    status_res = await db.execute(text("""
        SELECT COALESCE(status, 'NULL') as status, COUNT(*) as cnt
        FROM historical_matches
        GROUP BY status
        ORDER BY cnt DESC
    """))
    status_dist = {r.status: r.cnt for r in status_res.fetchall()}

    return {
        "matches": {
            "total": total_matches,
            "con_user_id_a": linked_a,
            "con_user_id_b": linked_b,
            "ambos_vinculados": both_linked,
            "persona_a_sin_vincular": orphan_a,
            "persona_b_sin_vincular": orphan_b,
            "porcentaje_vinculado": round((both_linked / total_matches * 100), 1) if total_matches > 0 else 0
        },
        "personas_no_registradas": unregistered_names,
        "usuarios": {
            "total": total_users,
            "con_codigo_dl": with_code,
            "con_cedula": with_cedula
        },
        "distribucion_status": status_dist
    }


from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File

# ─── PLAN SYNC FROM EXCEL ───────────────────────────────────────────────────

@router.post("/sync-plans")
async def sync_plans_from_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("importar", "use"))
):
    """
    Sincroniza automáticamente plan_tier en profiles asociando por correo y nombre
    a partir del archivo Excel cargado (pestaña 'Clients plans').
    """
    import openpyxl
    import io

    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        if "Clients plans" not in wb.sheetnames:

            return {"status": "error", "message": "Pestaña 'Clients plans' no existe en el Excel cargado"}

        # Ensure is_difficult and difficult_notes columns exist in profiles
        await db.execute(text("ALTER TABLE profiles ADD COLUMN IF NOT EXISTS is_difficult BOOLEAN DEFAULT false;"))
        await db.execute(text("ALTER TABLE profiles ADD COLUMN IF NOT EXISTS difficult_notes TEXT;"))
        await db.commit()

        ws = wb["Clients plans"]



        plans_config = [
            ("VIP 195k", 0, 1),
            ("Premium 150k", 2, 3),
            ("Estándar Plus 98k", 4, 5),
            ("Estándar 65k (2 citas)", 6, 7),
            ("Estándar 65k (1 cita)", 8, 9),
            ("Básico 40k", 10, 11)
        ]

        updated = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            for plan_name, name_idx, email_idx in plans_config:
                name_val = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] else ""
                email_val = str(row[email_idx]).strip() if email_idx < len(row) and row[email_idx] else ""
                
                if name_val or email_val:
                    res = await db.execute(text("""
                        UPDATE profiles p
                        SET plan_tier = :plan
                        FROM users u
                        WHERE p.user_id = u.id
                          AND (
                            (:email != '' AND lower(trim(COALESCE(u.email,''))) = lower(trim(:email)))
                            OR (:name != '' AND unaccent(lower(trim(COALESCE(u.name,'')))) = unaccent(lower(trim(:name))))
                          )
                    """), {"plan": plan_name, "email": email_val, "name": name_val})
                    updated += res.rowcount

        # ─── PARSE AGE, CITY & PLAN FROM PROFILES COL 4 (CIUDAD FREEFORM) ─
        import re
        ages_updated = 0
        if "PROFILES" in wb.sheetnames:
            ws_prof = wb["PROFILES"]
            for row in ws_prof.iter_rows(min_row=2, values_only=True):
                if not row or not row[1]: continue
                p_name = str(row[1]).strip()
                col4_text = str(row[4] or "").strip()
                col8_age = row[8] if len(row) > 8 else None

                # Extract age number if present in col4 or col8
                extracted_age = None
                if col8_age:
                    try: extracted_age = int(float(str(col8_age)))
                    except: pass
                
                if not extracted_age and col4_text:
                    # Match patterns like "27", "29", "49 años", "34 años"
                    age_match = re.search(r'\b(1[89]|[2-8][0-9])\b', col4_text)
                    if age_match:
                        extracted_age = int(age_match.group(1))

                # Extract city
                extracted_city = None
                if col4_text:
                    c_low = col4_text.lower()
                    if 'bog' in c_low or 'bgta' in c_low: extracted_city = 'Bogotá'
                    elif 'med' in c_low or 'mde' in c_low: extracted_city = 'Medellín'
                    elif 'cali' in c_low: extracted_city = 'Cali'
                    elif 'barranq' in c_low or 'bqui' in c_low: extracted_city = 'Barranquilla'
                    elif 'buca' in c_low: extracted_city = 'Bucaramanga'

                # Extract VIP if written in Col 4
                is_vip = 'vip' in col4_text.lower() if col4_text else False

                # Extract Col 2 Date
                col2_date = row[2] if len(row) > 2 else None
                extracted_date = None
                if col2_date:
                    if hasattr(col2_date, 'strftime'):
                        extracted_date = col2_date
                    else:
                        try:
                            extracted_date = datetime.strptime(str(col2_date)[:10], "%Y-%m-%d")
                        except: pass

                if extracted_age or extracted_city or is_vip or extracted_date:
                    updates = []
                    u_updates = []
                    params = {"person": p_name}
                    if extracted_age:
                        updates.append("age = :age")
                        params["age"] = extracted_age

                    if extracted_city:
                        updates.append("city = :city")
                        params["city"] = extracted_city
                    if is_vip:
                        updates.append("plan_tier = 'VIP 195k'")

                    if updates:
                        sql_up = f"""
                            UPDATE profiles p
                            SET {", ".join(updates)}
                            FROM users u
                            WHERE p.user_id = u.id
                              AND unaccent(lower(trim(u.name))) = unaccent(lower(trim(:person)))
                        """
                        r_up = await db.execute(text(sql_up), params)
                        if extracted_age: ages_updated += r_up.rowcount

                    if extracted_date:
                        await db.execute(text("""
                            UPDATE users
                            SET created_at = :cdate
                            WHERE unaccent(lower(trim(name))) = unaccent(lower(trim(:person)))
                        """), {"cdate": extracted_date, "person": p_name})

            await db.commit()



        # ─── ENRICH CITY AND RESPONSABLE FROM MATCHES TABS ─────────────────

        matches_sheets = [s for s in wb.sheetnames if 'MATCHES' in s and s not in ('MISSING MATCHES', 'TROUBLE MATCHES')]
        cities_updated = 0
        resp_updated = 0

        for sname in matches_sheets:
            ws_m = wb[sname]
            try:
                row1 = [str(c).upper() if c else "" for c in next(ws_m.iter_rows(min_row=1, max_row=1, values_only=True))]
            except StopIteration:
                continue

            city_idx = -1
            pa_idx = -1
            pb_idx = -1
            for idx, h in enumerate(row1):
                if 'CITY' in h or 'CIUDAD' in h: city_idx = idx
                if 'PERSON A' in h: pa_idx = idx
                if 'PERSON B' in h: pb_idx = idx

            psych_name = sname.replace('MATCHES', '').strip()

            for row in ws_m.iter_rows(min_row=2, values_only=True):
                if not row: continue
                city_val = str(row[city_idx]).strip() if city_idx >= 0 and city_idx < len(row) and row[city_idx] else ""
                pa_val = str(row[pa_idx]).strip() if pa_idx >= 0 and pa_idx < len(row) and row[pa_idx] else ""
                pb_val = str(row[pb_idx]).strip() if pb_idx >= 0 and pb_idx < len(row) and row[pb_idx] else ""

                normalized_city = ""
                if city_val and city_val.lower() not in ('none', 'null', 'city', 'ciudad', '1 o 2'):
                    c_clean = city_val.lower()
                    if 'bog' in c_clean or 'bgta' in c_clean: normalized_city = 'Bogotá'
                    elif 'med' in c_clean or 'mde' in c_clean: normalized_city = 'Medellín'
                    elif 'cali' in c_clean: normalized_city = 'Cali'
                    elif 'barranq' in c_clean: normalized_city = 'Barranquilla'
                    else: normalized_city = city_val.title()

                for p_person in (pa_val, pb_val):
                    if not p_person or len(p_person) < 3 or p_person.lower() in ('person a', 'person b', 'none', 'null'):
                        continue

                    # Update city if found
                    if normalized_city:
                        r_city = await db.execute(text("""
                            UPDATE profiles p
                            SET city = :city
                            FROM users u
                            WHERE p.user_id = u.id
                              AND unaccent(lower(trim(u.name))) = unaccent(lower(trim(:person)))
                        """), {"city": normalized_city, "person": p_person})
                        cities_updated += r_city.rowcount

                    # Update responsable from tab name
                    if psych_name:
                        r_resp = await db.execute(text("""
                            UPDATE profiles p
                            SET responsable = :psych
                            FROM users u
                            WHERE p.user_id = u.id
                              AND unaccent(lower(trim(u.name))) = unaccent(lower(trim(:person)))
                        """), {"psych": psych_name, "person": p_person})
                        resp_updated += r_resp.rowcount


        # ─── ENRICH SEARCH PREFERENCES FROM PREFERENCES SHEET ──────────────
        pref_updated = 0
        if "PREFERENCES" in wb.sheetnames:
            ws_p = wb["PREFERENCES"]
            for r in ws_p.iter_rows(min_row=2, values_only=True):
                if not r or not r[0]: continue
                try:
                    pid = int(float(str(r[0])))
                except:
                    continue

                intent = str(r[1] or '').strip() if len(r) > 1 else ''
                pref_gender = str(r[2] or '').strip() if len(r) > 2 else ''
                age_min = str(r[3] or '').strip() if len(r) > 3 else ''
                age_max = str(r[4] or '').strip() if len(r) > 4 else ''
                values = str(r[6] or '').strip() if len(r) > 6 else ''
                notes = str(r[13] or '').strip() if len(r) > 13 else ''
                red_flags = str(r[14] or '').strip() if len(r) > 14 else ''
                green_flags = str(r[15] or '').strip() if len(r) > 15 else ''

                pref_dict = {}
                if intent: pref_dict["intent"] = intent
                if pref_gender: pref_dict["pref_gender"] = pref_gender
                if age_min or age_max: pref_dict["preferred_age_range"] = f"{age_min}-{age_max}"
                if values: pref_dict["must_have_values"] = values
                if notes: pref_dict["looks_notes"] = notes
                if red_flags: pref_dict["red_flags"] = red_flags
                if green_flags: pref_dict["green_flags"] = green_flags

                if pref_dict:
                    import json
                    pref_json = json.dumps(pref_dict)
                    r_pref = await db.execute(text("""
                        UPDATE profiles
                        SET search_preferences = CAST(:pref AS jsonb)
                        WHERE user_id = :uid
                          AND (search_preferences IS NULL OR search_preferences = '{}'::jsonb)
                    """), {"pref": pref_json, "uid": pid})
                    pref_updated += r_pref.rowcount

        # ─── ENRICH PERSONAS DIFICILES FROM PERSONAS DIFICILES SHEET ─────
        diff_updated = 0
        if "PERSONAS DÍFICILES" in wb.sheetnames:
            ws_d = wb["PERSONAS DÍFICILES"]
            for r in ws_d.iter_rows(min_row=2, values_only=True):
                if not r or not r[0]: continue
                p_name = str(r[0]).strip()
                obs = str(r[2] or '').strip() if len(r) > 2 else ''
                status_d = str(r[3] or '').strip() if len(r) > 3 else ''

                if p_name and len(p_name) > 2:
                    r_diff = await db.execute(text("""
                        UPDATE profiles p
                        SET is_difficult = true,
                            difficult_notes = COALESCE(:obs, '')
                        FROM users u
                        WHERE p.user_id = u.id
                          AND unaccent(lower(trim(u.name))) = unaccent(lower(trim(:person)))
                    """), {"obs": f"{obs} {status_d}".strip(), "person": p_name})
                    diff_updated += r_diff.rowcount

        # General cleanup of remaining legacy yes/no city flags
        await db.execute(text("""
            UPDATE profiles
            SET city = 'Bogotá'
            WHERE (city IS NULL OR city IN ('yes', 'no', 'no ', 'yes ', ''))
              AND bio_notes ILIKE '%bogot%';
        """))

        await db.commit()

        # Resumen post-sync
        res_dist = await db.execute(text("SELECT COALESCE(plan_tier, 'Sin Plan') as tier, COUNT(*) as cnt FROM profiles GROUP BY plan_tier ORDER BY cnt DESC"))
        distribucion = {r.tier: r.cnt for r in res_dist.fetchall()}

        res_cities = await db.execute(text("SELECT COALESCE(city, 'Sin Ciudad') as city_name, COUNT(*) as cnt FROM profiles GROUP BY city ORDER BY cnt DESC LIMIT 10"))
        dist_ciudades = {r.city_name: r.cnt for r in res_cities.fetchall()}

        return {
            "status": "success",
            "registros_actualizados": updated,
            "ciudades_enriquecidas": cities_updated,
            "psicologas_enriquecidas": resp_updated,
            "personas_dificiles_etiquetadas": diff_updated,
            "distribucion_actual": distribucion,
            "distribucion_ciudades": dist_ciudades
        }
    except Exception as e:
        return {"status": "error", "detail": str(e)}




# ─── STATS ────────────────────────────────────────────────────────────────────


@router.get("/stats")
async def get_stats(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("dashboard", "view"))
):
    """Dashboard KPI summary."""
    total_users = (await db.execute(text("SELECT COUNT(*) FROM users WHERE merged_into_id IS NULL"))).scalar() or 0

    events_this_month = (await db.execute(text("""
        SELECT COUNT(*) FROM events
        WHERE date >= date_trunc('month', NOW())
          AND date < date_trunc('month', NOW()) + INTERVAL '1 month'
    """))).scalar() or 0

    active_matches = (await db.execute(text(
        "SELECT (SELECT COUNT(*) FROM match_requests WHERE status = 'pending') + (SELECT COUNT(*) FROM historical_matches)"
    ))).scalar() or 0

    avg_sat_row = (await db.execute(text(
        "SELECT ROUND(AVG(satisfaccion)::numeric, 1) FROM post_event_feedback WHERE satisfaccion IS NOT NULL"
    ))).scalar()
    avg_satisfaction = float(avg_sat_row) if avg_sat_row else None

    # Weekly growth: new users per week for last 8 weeks (excluding merged users)
    weekly_rows = (await db.execute(text("""
        SELECT
            EXTRACT(WEEK FROM created_at) AS wk,
            COUNT(*) AS cnt
        FROM users
        WHERE merged_into_id IS NULL AND created_at >= NOW() - INTERVAL '8 weeks'
        GROUP BY wk
        ORDER BY wk
    """))).fetchall()
    weekly_growth = [int(r.cnt) for r in weekly_rows]

    # Alert-specific metrics
    active_debts = (await db.execute(text(
        "SELECT COUNT(*) FROM accounts_receivable WHERE status = 'pending' AND due_date < CURRENT_DATE"
    ))).scalar() or 0

    pending_payrolls = (await db.execute(text(
        "SELECT COUNT(*) FROM payroll_runs WHERE status = 'draft'"
    ))).scalar() or 0

    critical_events = (await db.execute(text("""
        SELECT COUNT(*) FROM (
            SELECT e.id, e.capacity, COUNT(ea.user_id) as attendees_count
            FROM events e
            LEFT JOIN event_attendees ea ON ea.event_id = e.id AND ea.status IN ('confirmed', 'attended')
            WHERE e.date > NOW()
            GROUP BY e.id, e.capacity
        ) sub WHERE attendees_count >= (capacity * 0.85) AND capacity > 0
    """))).scalar() or 0

    return {
        "total_users": total_users,
        "events_this_month": events_this_month,
        "active_matches": active_matches,
        "avg_satisfaction": avg_satisfaction,
        "weekly_growth": weekly_growth,
        "active_debts": active_debts,
        "pending_payrolls": pending_payrolls,
        "critical_events": critical_events,
    }


@router.get("/today-summary")
async def get_today_summary(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("dashboard", "view"))
):
    """Consolidated operational summary of the next 48 hours."""
    now = datetime.now()
    month = now.month
    year = now.year

    # 1. Upcoming events (next 48 hours)
    event_rows = (await db.execute(text("""
        SELECT e.id, e.name, e.date, e.capacity,
               (SELECT COUNT(*) FROM event_attendees WHERE event_id = e.id AND status IN ('confirmed', 'attended')) as confirmed_attendees
        FROM events e
        WHERE e.date >= NOW() AND e.date <= NOW() + INTERVAL '48 hours'
        ORDER BY e.date ASC
    """))).fetchall()

    upcoming_events = []
    for r in event_rows:
        capacity = r.capacity or 100
        confirmed = r.confirmed_attendees or 0
        pct = (confirmed / capacity * 100.0) if capacity > 0 else 0.0
        
        # Hours until event
        time_diff = r.date - now
        hours_until = max(0.0, time_diff.total_seconds() / 3600.0)
        
        status = "critical" if pct > 85 else "warning" if pct > 70 else "normal"
        
        upcoming_events.append({
            "id": r.id,
            "name": r.name,
            "date": r.date.isoformat(),
            "hours_until": round(hours_until, 1),
            "capacity": capacity,
            "confirmed_attendees": confirmed,
            "occupancy_pct": round(pct, 1),
            "status": status
        })

    # 2. Overdue payments (cartera vencida)
    overdue_rows = (await db.execute(text("""
        SELECT ar.id, ar.user_id, ar.amount, ar.due_date, u.name as user_name,
               (CURRENT_DATE - ar.due_date) as days_overdue
        FROM accounts_receivable ar
        JOIN users u ON u.id = ar.user_id
        WHERE ar.due_date < CURRENT_DATE AND ar.status = 'pending'
        ORDER BY days_overdue DESC
        LIMIT 10
    """))).fetchall()

    overdue_payments = [{
        "id": str(r.id),
        "user_id": r.user_id,
        "user_name": r.user_name,
        "amount": float(r.amount),
        "days_overdue": max(0, int(r.days_overdue))
    } for r in overdue_rows]

    # 3. Pending payroll for current month
    payroll_draft = (await db.execute(text("""
        SELECT COUNT(*) as cnt, COALESCE(SUM(total_paid), 0) as total 
        FROM payroll_runs 
        WHERE period_month = :month AND period_year = :year AND status = 'draft'
        GROUP BY id
    """), {"month": month, "year": year})).fetchone()

    if payroll_draft:
        pending_payroll = {
            "has_pending": True,
            "period": f"{month}/{year}",
            "employees_count": int(payroll_draft.cnt),
            "estimated_total": float(payroll_draft.total)
        }
    else:
        # Generate calculation estimate
        emp_stats = (await db.execute(text("""
            SELECT COUNT(*) as count, COALESCE(SUM(base_salary), 0) as base_sum
            FROM employees
            WHERE status = 'active'
        """))).fetchone()
        
        pending_commissions = (await db.execute(text("""
            SELECT COALESCE(SUM(amount), 0)
            FROM employee_event_commissions
            WHERE status = 'pending'
        """))).scalar() or 0.0
        
        has_active_employees = emp_stats and emp_stats.count > 0
        
        pending_payroll = {
            "has_pending": has_active_employees,
            "period": f"{month}/{year}",
            "employees_count": int(emp_stats.count) if emp_stats else 0,
            "estimated_total": float(emp_stats.base_sum or 0.0) + float(pending_commissions)
        }

    # 4. Recent incidents (Mejora 2)
    incident_rows = (await db.execute(text("""
        SELECT ei.id, ei.event_id, ei.category, ei.severity, ei.description, ei.resolved, ei.created_at,
               ev.name as event_name, emp.full_name as reported_by_name
        FROM event_incidents ei
        LEFT JOIN events ev ON ev.id = ei.event_id
        LEFT JOIN employees emp ON emp.id = ei.reported_by
        ORDER BY ei.created_at DESC
        LIMIT 10
    """))).fetchall()

    recent_incidents = [{
        "id": str(r.id),
        "event_id": r.event_id,
        "event_name": r.event_name,
        "reported_by_name": r.reported_by_name or "Sistema",
        "category": r.category,
        "severity": r.severity,
        "description": r.description,
        "resolved": r.resolved,
        "created_at": r.created_at.isoformat()
    } for r in incident_rows]

    return {
        "date": now.date().isoformat(),
        "upcoming_events": upcoming_events,
        "overdue_payments": overdue_payments,
        "pending_payroll": pending_payroll,
        "low_stock_alerts": [],
        "recent_incidents": recent_incidents
    }


# ─── USERS / CLIENTS ──────────────────────────────────────────────────────────

@router.get("/users")
async def get_users(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    responsable: Optional[str] = Query(None),
    has_notes: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    has_matches: Optional[str] = Query(None),
    plan_tier: Optional[str] = Query(None),
    is_difficult: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):
    """Paginated client list with rich backend SQL filters."""
    offset = (page - 1) * limit

    where_clauses = ["u.merged_into_id IS NULL"]
    params: dict = {"limit": limit, "offset": offset}

    if search:
        where_clauses.append("(unaccent(u.name) ILIKE unaccent(:search) OR u.email ILIKE :search OR u.phone ILIKE :search OR COALESCE(u.client_code,'') ILIKE :search OR COALESCE(u.id_number,'') ILIKE :search OR unaccent(COALESCE(p.occupation, '')) ILIKE unaccent(:search) OR unaccent(COALESCE(p.city, '')) ILIKE unaccent(:search))")
        params["search"] = f"%{search}%"

    if responsable and responsable != "all":
        r_clean = responsable.strip().upper()
        aliases = [r_clean]
        if "SILVI" in r_clean or "SILVIA" in r_clean:
            aliases = ["SILVI", "SILVIA", "SILV"]
        elif "STEFF" in r_clean or "STEPH" in r_clean:
            aliases = ["STEFFY", "STEFF", "STEPHANIE"]
        elif "PAULA" in r_clean or "MAPE" in r_clean:
            aliases = ["MARÍA PAULA", "MARIA PAULA", "MAPE", "PAULA"]
        elif "MANU" in r_clean:
            aliases = ["MANU", "MANUELA"]

        alias_conds = " OR ".join([f"unaccent(COALESCE(p.responsable, '')) ILIKE '%{a}%'" for a in aliases])
        where_clauses.append(f"({alias_conds})")

    if is_difficult == "difficult_only":
        where_clauses.append("COALESCE(p.is_difficult, false) = true")


    if has_notes == "with_notes":
        where_clauses.append("p.bio_notes IS NOT NULL AND length(trim(p.bio_notes)) > 2")
    elif has_notes == "without_notes":
        where_clauses.append("(p.bio_notes IS NULL OR length(trim(p.bio_notes)) <= 2)")

    if city and city != "all":
        city_clean = city.strip()
        if "bogot" in city_clean.lower():
            where_clauses.append("(p.city ILIKE '%bogot%' OR p.city ILIKE '%bogot%' OR unaccent(COALESCE(p.city, '')) ILIKE '%bogot%' OR p.bio_notes ILIKE '%bogot%')")
        elif "medell" in city_clean.lower():
            where_clauses.append("(p.city ILIKE '%medell%' OR unaccent(COALESCE(p.city, '')) ILIKE '%medell%' OR p.bio_notes ILIKE '%medell%')")
        else:
            where_clauses.append("(unaccent(COALESCE(p.city, '')) ILIKE unaccent(:city) OR unaccent(COALESCE(p.bio_notes, '')) ILIKE unaccent(:city) OR unaccent(COALESCE(CAST(p.search_preferences AS text), '')) ILIKE unaccent(:city))")
            params["city"] = f"%{city_clean}%"


    if plan_tier and plan_tier != "all":
        if plan_tier == "sin_plan":
            where_clauses.append("(p.plan_tier IS NULL OR trim(p.plan_tier) = '')")
        else:
            where_clauses.append("UPPER(COALESCE(p.plan_tier, '')) LIKE UPPER(:plan_tier)")
            params["plan_tier"] = f"%{plan_tier}%"


    if has_matches == "with_matches":
        where_clauses.append("""unaccent(lower(u.name)) IN (
            SELECT DISTINCT unaccent(lower(person_a)) FROM historical_matches WHERE person_a IS NOT NULL
            UNION
            SELECT DISTINCT unaccent(lower(person_b)) FROM historical_matches WHERE person_b IS NOT NULL
        )""")
    elif has_matches == "without_matches":
        where_clauses.append("""unaccent(lower(u.name)) NOT IN (
            SELECT DISTINCT unaccent(lower(person_a)) FROM historical_matches WHERE person_a IS NOT NULL
            UNION
            SELECT DISTINCT unaccent(lower(person_b)) FROM historical_matches WHERE person_b IS NOT NULL
        )""")

    where_str = " AND ".join(where_clauses)

    total = (await db.execute(text(f"""
        SELECT COUNT(*)
        FROM users u
        LEFT JOIN profiles p ON p.user_id = u.id
        WHERE {where_str}
    """), params)).scalar() or 0

    rows = (await db.execute(text(f"""
        SELECT
            u.id, u.phone, u.name, u.created_at,
            u.client_code, u.id_number,
            p.user_id AS profile_user_id, p.ocean, p.apego, p.motivacion, p.rol_social,
            p.energia_social, p.momento_vital, p.intereses, p.valores,
            p.city, p.occupation, p.education, p.religion, p.love_language,
            p.bio_notes, p.lifestyle, p.responsable, p.estatura, p.age, p.plan_tier, p.search_preferences,
            COALESCE(p.is_difficult, false) AS is_difficult, p.difficult_notes,
            COALESCE(hm_count.cnt, 0) AS total_matches
        FROM users u
        LEFT JOIN profiles p ON p.user_id = u.id
        LEFT JOIN (
            SELECT uid, COUNT(*) as cnt FROM (
                SELECT user_id_a AS uid FROM historical_matches WHERE user_id_a IS NOT NULL
                UNION ALL
                SELECT user_id_b AS uid FROM historical_matches WHERE user_id_b IS NOT NULL
            ) sub GROUP BY uid
        ) hm_count ON hm_count.uid = u.id
        WHERE {where_str}
        ORDER BY u.id DESC
        LIMIT :limit OFFSET :offset
    """), params)).fetchall()


    users_list = []
    for r in rows:
        ocean = json.loads(r.ocean) if isinstance(r.ocean, str) else (r.ocean or {})
        lifestyle = json.loads(r.lifestyle) if isinstance(r.lifestyle, str) else (r.lifestyle or {})
        search_prefs = json.loads(r.search_preferences) if isinstance(r.search_preferences, str) else (r.search_preferences or {})
        intereses = json.loads(r.intereses) if isinstance(r.intereses, str) else (r.intereses or [])
        valores = json.loads(r.valores) if isinstance(r.valores, str) else (r.valores or [])

        client_code = r.client_code or f"DL-{str(r.id).zfill(4)}"
        users_list.append({
            "id": r.id,
            "phone": r.phone,
            "name": r.name or "Sin nombre",
            "client_code": client_code,
            "id_number": r.id_number,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "has_profile": r.profile_user_id is not None,
            "city": r.city,
            "occupation": r.occupation,
            "education": r.education,
            "religion": r.religion,
            "love_language": r.love_language,
            "bio_notes": r.bio_notes,
            "responsable": r.responsable,
            "estatura": r.estatura,
            "age": r.age,
            "plan_tier": r.plan_tier,
            "is_difficult": r.is_difficult,
            "difficult_notes": r.difficult_notes,
            "total_matches": r.total_matches or 0,
            "profile": {


                "city": r.city or "Bogotá",
                "age": r.age or 28,
                "occupation": r.occupation or "No especificada",
                "education": r.education or "No especificada",
                "religion": r.religion or "No especificada",
                "love_language": r.love_language or "No especificado",
                "bio_notes": r.bio_notes or "",
                "responsable": r.responsable or "SILVI",
                "estatura": r.estatura or "No especificada",
                "plan_tier": r.plan_tier or "Estándar",
                "motivacion": r.motivacion or "Conexión profunda",
                "apego": r.apego or "Seguro",
                "rol_social": r.rol_social or "Equilibrado",
                "energia_social": r.energia_social or "Ambivertido",
                "ocean": ocean if (ocean and isinstance(ocean, dict) and len(ocean) > 0) else None,
                "lifestyle": lifestyle if lifestyle else None,
                "search_preferences": search_prefs if search_prefs else None,

                "intereses": intereses,
                "valores": valores,
            }
        })


    if search:
        existing_names_lower = {u["name"].strip().lower() for u in users_list if u.get("name")}
        hm_rows = (await db.execute(text("""
            SELECT DISTINCT person_a AS name, city, matchmaker AS responsable, id
            FROM historical_matches
            WHERE unaccent(lower(person_a)) ILIKE unaccent(lower(:search))
            UNION
            SELECT DISTINCT person_b AS name, city, matchmaker AS responsable, id
            FROM historical_matches
            WHERE unaccent(lower(person_b)) ILIKE unaccent(lower(:search))
            LIMIT :limit
        """), {"search": f"%{search}%", "limit": limit})).fetchall()
        
        for r in hm_rows:
            clean_name = (r.name or '').strip()
            if clean_name and clean_name.lower() not in existing_names_lower:
                existing_names_lower.add(clean_name.lower())
                users_list.append({
                    "id": 9000 + int(r.id),
                    "phone": "Importado desde Excel",
                    "name": clean_name,
                    "created_at": datetime.now().isoformat(),
                    "has_profile": True,
                    "city": r.city or "Bogotá",
                    "occupation": "Cliente Histórico Matchmaking",
                    "responsable": r.responsable or "SILVI",
                    "motivacion": "conexion_profunda",
                    "age": 28,
                    "profile": {
                        "ocean": {"apertura": 0.85, "responsabilidad": 0.8, "extroversion": 0.75, "amabilidad": 0.9, "neuroticismo": 0.2},
                        "apego": "Seguro",
                        "motivacion": "conexion_profunda",
                        "city": r.city or "Bogotá"
                    }
                })

    return {"users": users_list, "total": len(users_list) if search else total, "page": page, "pages": math.ceil((len(users_list) if search else total) / limit)}


@router.get("/users/{user_id}")
async def get_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):
    """Full profile of a single user."""
    row = (await db.execute(text("""
        SELECT
            u.id, u.phone, u.name, u.created_at,
            u.client_code, u.id_number,
            p.ocean, p.apego, p.motivacion, p.rol_social,
            p.energia_social, p.momento_vital, p.intereses, p.valores, p.raw_answers
        FROM users u
        LEFT JOIN profiles p ON p.user_id = u.id
        WHERE u.id = :uid
    """), {"uid": user_id})).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # Events attended
    events = (await db.execute(text("""
        SELECT e.id, e.name, e.date, ea.status
        FROM event_attendees ea
        JOIN events e ON e.id = ea.event_id
        WHERE ea.user_id = :uid
        ORDER BY e.date DESC
        LIMIT 10
    """), {"uid": user_id})).fetchall()

    client_code = row.client_code or f"DL-{str(row.id).zfill(4)}"
    return {
        "id": row.id,
        "phone": row.phone,
        "name": row.name,
        "client_code": client_code,
        "id_number": row.id_number,
        "created_at": row.created_at.isoformat() if row.created_at else None,
        "profile": {
            "ocean": row.ocean,
            "apego": row.apego,
            "motivacion": row.motivacion,
            "rol_social": row.rol_social,
            "energia_social": row.energia_social,
            "momento_vital": row.momento_vital,
            "intereses": row.intereses,
            "valores": row.valores,
        },
        "events": [{"id": e.id, "name": e.name, "date": e.date.isoformat(), "status": e.status} for e in events]
    }


@router.put("/users/{user_id}")
async def update_user(
    user_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "edit"))
):
    """Update basic user fields."""
    allowed = {k: v for k, v in data.items() if k in ("name", "phone", "id_number")}
    if not allowed:
        raise HTTPException(status_code=400, detail="No hay campos válidos para actualizar")

    # If id_number is provided, check for duplicates before saving
    if "id_number" in allowed and allowed["id_number"]:
        dup = (await db.execute(text(
            "SELECT id FROM users WHERE id_number = :idn AND id != :uid"
        ), {"idn": allowed["id_number"].strip(), "uid": user_id})).fetchone()
        if dup:
            raise HTTPException(
                status_code=409,
                detail=f"Ya existe otro usuario registrado con esa cédula/documento (ID: DL-{str(dup.id).zfill(4)}). Verifica que no sea un duplicado."
            )
        allowed["id_number"] = allowed["id_number"].strip() or None

    set_clause = ", ".join(f"{k} = :{k}" for k in allowed)
    allowed["uid"] = user_id
    await db.execute(text(f"UPDATE users SET {set_clause} WHERE id = :uid"), allowed)
    await db.commit()
    return {"ok": True}


@router.get("/users/{user_id}/match-analysis")
async def analyze_user_matchmaking_viability(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("matching", "view"))
):
    """
    Algoritmo de diagnóstico clínico de IA para analizar la viabilidad de matches de un cliente.
    Identifica razones cuando un usuario no tiene matches (perfil incompleto, falta de candidatas/os en la ciudad, etc.).
    """
    usr_res = await db.execute(text("""
        SELECT u.id, u.name, u.phone, u.client_code, u.id_number,
               p.gender, p.city, p.age, p.occupation, p.photo_url, p.bio_notes,
               p.responsable, p.search_preferences, p.motivacion
        FROM users u
        LEFT JOIN profiles p ON p.user_id = u.id
        WHERE u.id = :uid
    """), {"uid": user_id})
    usr = usr_res.fetchone()

    if not usr:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # 1. Contar citas reales en historial
    matches_count_res = await db.execute(text("""
        SELECT COUNT(*) FROM historical_matches
        WHERE user_id_a = :uid OR user_id_b = :uid
           OR unaccent(lower(trim(person_a))) = unaccent(lower(trim(:name)))
           OR unaccent(lower(trim(person_b))) = unaccent(lower(trim(:name)))
    """), {"uid": user_id, "name": (usr.name or "").strip()})
    real_matches_count = matches_count_res.scalar() or 0

    # 2. Evaluar completitud del perfil
    missing_fields = []
    if not usr.name: missing_fields.append("Nombre")
    if not usr.phone: missing_fields.append("Teléfono")
    if not usr.gender: missing_fields.append("Género")
    if not usr.city: missing_fields.append("Ciudad")
    if not usr.age: missing_fields.append("Edad")
    if not usr.photo_url: missing_fields.append("Fotografía de Lookbook")
    if not usr.responsable: missing_fields.append("Psicóloga asignada")

    completeness = round(((7 - len(missing_fields)) / 7) * 100)

    # 3. Determinar género buscado
    user_gender = (usr.gender or "").strip().lower()
    target_gender = "femenino" if user_gender in ["masculino", "hombre", "m"] else ("masculino" if user_gender in ["femenino", "mujer", "f"] else "todos")

    user_city = (usr.city or "Bogotá").strip()
    user_age = usr.age or 30

    # 3. VERIFICACIÓN DE EVALUACIÓN POST-CITA OBLIGATORIA
    pending_feedback_res = await db.execute(text("""
        SELECT COUNT(*) FROM historical_matches hm
        WHERE (
            ((hm.user_id_a = :uid OR unaccent(lower(trim(hm.person_a))) = unaccent(lower(trim(:uname)))) AND hm.feedback_completed_a IS FALSE)
            OR
            ((hm.user_id_b = :uid OR unaccent(lower(trim(hm.person_b))) = unaccent(lower(trim(:uname)))) AND hm.feedback_completed_b IS FALSE)
        )
        AND (hm.status ILIKE '%APROBADO%' OR hm.status ILIKE '%REALIZADA%' OR hm.status ILIKE '%HECHO%' OR hm.status ILIKE '%CONFIRMADO%')
    """), {"uid": user_id, "uname": user_name})
    pending_feedback_count = pending_feedback_res.scalar() or 0

    if pending_feedback_count > 0:
        return {
            "viability": {
                "status": "BLOCKED",
                "badge": "⛔ BLOQUEADO: EVALUACIÓN POST-CITA PENDIENTE",
                "title": f"⛔ {user_name} tiene {pending_feedback_count} evaluación(es) post-cita pendiente(s) por responder.",
                "reasons": [
                    "Requisito obligatorio: El cliente debe responder la encuesta enviada por correo electrónico evaluando la cita, el lugar y la persona.",
                    "Su perfil permanecerá bloqueado y no recibirá nuevas propuestas de matchmaking hasta enviar su retroalimentación."
                ],
                "recommended_action": f"Re-enviar encuesta por correo electrónico a {user_name} para desbloquear su perfil."
            },
            "candidate_pool": {"city": user_city, "count": 0}
        }

    # 4. Consultar pool de candidatos compatibles
    pool_query = """
        SELECT COUNT(*) FROM users u
        JOIN profiles p ON p.user_id = u.id
        WHERE u.id != :uid
    """

    params = {"uid": user_id, "city": f"%{user_city}%"}

    if target_gender == "femenino":
        pool_query += " AND lower(COALESCE(p.gender, '')) IN ('femenino', 'mujer', 'f')"
    elif target_gender == "masculino":
        pool_query += " AND lower(COALESCE(p.gender, '')) IN ('masculino', 'hombre', 'm')"

    # 4b. Análisis de tendencias del pool de candidatos en la misma ciudad
    cands_traits_res = await db.execute(text(f"""
        SELECT p.motivacion, p.apego, p.occupation, p.intereses
        FROM users u
        JOIN profiles p ON p.user_id = u.id
        WHERE u.id != :uid
          AND p.gender IS NOT NULL AND length(trim(p.gender)) > 0
          AND p.city IS NOT NULL AND unaccent(lower(trim(p.city))) = unaccent(lower(trim(:city)))
          AND lower(trim(p.gender)) IN ({ "'femenino', 'mujer', 'f'" if target_gender == "femenino" else "'masculino', 'hombre', 'm'" if target_gender == "masculino" else "'femenino', 'masculino'" })
        LIMIT 30
    """), {"uid": user_id, "city": user_city})
    cands_traits = cands_traits_res.fetchall()

    common_motivations = {}
    common_apego = {}
    for r in cands_traits:
        if r.motivacion:
            mot = r.motivacion.replace('_', ' ').title()
            common_motivations[mot] = common_motivations.get(mot, 0) + 1
        if r.apego:
            ap = r.apego.title()
            common_apego[ap] = common_apego.get(ap, 0) + 1

    top_mot = max(common_motivations, key=common_motivations.get) if common_motivations else None
    top_ap = max(common_apego, key=common_apego.get) if common_apego else None

    total_target_gender = (await db.execute(text(pool_query), params)).scalar() or 0

    pool_city_query = pool_query + " AND (unaccent(lower(COALESCE(p.city, ''))) ILIKE unaccent(lower(:city)) OR p.city IS NULL)"
    city_compatible_candidates = (await db.execute(text(pool_city_query), params)).scalar() or 0

    # 5. Generar diagnóstico clínico y recomendación indagatoria para la psicóloga
    clinical_reasons = []
    recommended_action = ""

    if real_matches_count > 0:
        diagnosis_title = f"Cliente Activo con {real_matches_count} Citas Registradas"
        clinical_reasons.append(f"El usuario tiene {real_matches_count} propuestas/citas agendadas en su expediente histórico.")
        recommended_action = "Continuar seguimiento post-cita y recabar retroalimentación del cliente."
    else:
        diagnosis_title = "Sin Citas Agendadas Actuales"

        if len(missing_fields) > 0:
            clinical_reasons.append(f"Ficha clínica incompleta ({completeness}% completado). Campos faltantes: {', '.join(missing_fields)}.")
            recommended_action = f"Completar la información clínica faltante ({', '.join(missing_fields)}) para ingresar al proceso de matching."
        
        if city_compatible_candidates == 0:
            clinical_reasons.append(f"Sin candidatos con perfil 100% completo del género {target_gender.capitalize()} en {user_city}.")
            recommended_action = f"Ampliar la prospección clínica en {user_city} o invitar a completar fichas pendientes."
        else:
            inquiry_points = []
            if top_mot: inquiry_points.append(f"Motivación dominante en candidatas/os: '{top_mot}'")
            if top_ap: inquiry_points.append(f"Estilo de apego predominante: '{top_ap}'")
            inquiry_str = f" ({', '.join(inquiry_points)})" if inquiry_points else ""

            clinical_reasons.append(f"Existen {city_compatible_candidates} candidatos potenciales en {user_city} ({target_gender.capitalize()}) con perfil verificado.")
            recommended_action = f"Se recomienda a la psicóloga {usr.responsable or 'asignada'} indagar en la entrevista clínica sobre proyecto de vida y valores{inquiry_str} para filtrar de forma exacta con las {city_compatible_candidates} opciones de {user_city}."

    return {
        "user_id": usr.id,
        "client_code": usr.client_code,
        "user_name": usr.name,
        "city": user_city,
        "gender": usr.gender,
        "age": usr.age,
        "real_matches_count": real_matches_count,
        "completeness": completeness,
        "missing_fields": missing_fields,
        "pool_metrics": {
            "target_gender_searched": target_gender.capitalize(),
            "total_in_system": total_target_gender,
            "city_compatible": city_compatible_candidates,
            "top_motivation": top_mot,
            "top_apego": top_ap
        },
        "diagnostic": {
            "title": diagnosis_title,
            "reasons": clinical_reasons,
            "recommended_action": recommended_action
        }
    }




@router.get("/merged-users")
async def get_merged_users(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):
    """List all merged duplicate users for audit purposes."""
    rows = (await db.execute(text("""
        SELECT 
            u1.id AS secondary_id,
            u1.name AS secondary_name,
            u1.phone AS secondary_phone,
            u1.merged_at,
            u2.id AS primary_id,
            u2.name AS primary_name,
            u2.phone AS primary_phone
        FROM users u1
        JOIN users u2 ON u1.merged_into_id = u2.id
        WHERE u1.merged_into_id IS NOT NULL
        ORDER BY u1.merged_at DESC, u1.id DESC
    """))).fetchall()

    return [{
        "secondary_id": r.secondary_id,
        "secondary_name": r.secondary_name,
        "secondary_phone": r.secondary_phone,
        "merged_at": r.merged_at.isoformat() if r.merged_at else None,
        "primary_id": r.primary_id,
        "primary_name": r.primary_name,
        "primary_phone": r.primary_phone
    } for r in rows]


@router.get("/users/check-duplicate")
async def check_duplicate_user(
    id_number: Optional[str] = Query(None),
    phone: Optional[str] = Query(None),
    exclude_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):
    """Check if a cedula/document or phone already exists in the system."""
    result = {"exists": False, "conflicts": []}

    if id_number and id_number.strip():
        q = "SELECT id, name, client_code FROM users WHERE id_number = :idn"
        params = {"idn": id_number.strip()}
        if exclude_id:
            q += " AND id != :exc"
            params["exc"] = exclude_id
        dup = (await db.execute(text(q), params)).fetchone()
        if dup:
            cc = dup.client_code or f"DL-{str(dup.id).zfill(4)}"
            result["exists"] = True
            result["conflicts"].append({
                "field": "id_number",
                "message": f"La cédula ya está registrada para {dup.name or 'Sin nombre'} ({cc})",
                "existing_user": {"id": dup.id, "name": dup.name, "client_code": cc}
            })

    if phone and phone.strip():
        q = "SELECT id, name, client_code FROM users WHERE phone = :ph"
        params = {"ph": phone.strip()}
        if exclude_id:
            q += " AND id != :exc"
            params["exc"] = exclude_id
        dup = (await db.execute(text(q), params)).fetchone()
        if dup:
            cc = dup.client_code or f"DL-{str(dup.id).zfill(4)}"
            result["exists"] = True
            result["conflicts"].append({
                "field": "phone",
                "message": f"El teléfono ya está registrado para {dup.name or 'Sin nombre'} ({cc})",
                "existing_user": {"id": dup.id, "name": dup.name, "client_code": cc}
            })

    return result


@router.post("/users/{user_id}/assign-matchmaker")
async def assign_matchmaker(
    user_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "edit"))
):
    matchmaker = data.get("matchmaker") or data.get("responsable") or "SILVI"
    await db.execute(text("""
        UPDATE profiles SET responsable = :m, updated_at = NOW() WHERE user_id = :uid
    """), {"m": matchmaker, "uid": user_id})
    await db.commit()
    return {"message": f"Psicóloga asignada con éxito: {matchmaker}", "user_id": user_id, "responsable": matchmaker}


# ─── EVENTS ───────────────────────────────────────────────────────────────────

@router.get("/events")
async def get_events(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("eventos", "view"))
):
    """All events ordered by date descending."""
    rows = (await db.execute(text("""
        SELECT e.id, e.name, e.date, e.location, e.format, e.capacity, e.price, e.created_at,
               e.budget_income, e.budget_expenses,
               COUNT(ea.user_id) AS attendee_count
        FROM events e
        LEFT JOIN event_attendees ea ON ea.event_id = e.id
        GROUP BY e.id
        ORDER BY e.date DESC
    """))).fetchall()

    return {"events": [{
        "id": r.id, "name": r.name,
        "date": r.date.isoformat() if r.date else None,
        "location": r.location, "format": r.format,
        "capacity": r.capacity, "price": float(r.price) if r.price else None,
        "budget_income": float(r.budget_income) if r.budget_income is not None else None,
        "budget_expenses": float(r.budget_expenses) if r.budget_expenses is not None else None,
        "attendee_count": r.attendee_count,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    } for r in rows]}


@router.post("/events", status_code=201)
async def create_event(
    data: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("eventos", "create"))
):
    """Create a new event."""
    required = {"name", "date"}
    if not required.issubset(data.keys()):
        raise HTTPException(status_code=400, detail="'name' y 'date' son obligatorios")
    
    try:
        date_str = data["date"]
        if isinstance(date_str, str):
            date_str = date_str.replace(" ", "T").split(".")[0]
            event_date = datetime.fromisoformat(date_str)
        else:
            event_date = date_str
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Formato de fecha inválido: {str(e)}. Use AAAA-MM-DDTHH:MM")

    result = await db.execute(text("""
        INSERT INTO events (name, date, location, format, capacity, price, budget_income, budget_expenses)
        VALUES (:name, :date, :location, :format, :capacity, :price, :budget_income, :budget_expenses)
        RETURNING id
    """), {
        "name": data["name"],
        "date": event_date,
        "location": data.get("location"),
        "format": data.get("format"),
        "capacity": data.get("capacity"),
        "price": data.get("price"),
        "budget_income": data.get("budget_income"),
        "budget_expenses": data.get("budget_expenses"),
    })
    await db.commit()
    event_id = result.scalar()
    return {"id": event_id, "ok": True}


@router.put("/events/{event_id}")
async def update_event(
    event_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("eventos", "edit"))
):
    """Update an existing event."""
    exists = (await db.execute(text("SELECT 1 FROM events WHERE id = :id"), {"id": event_id})).scalar()
    if not exists:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
        
    try:
        date_str = data.get("date")
        event_date = None
        if date_str:
            if isinstance(date_str, str):
                date_str = date_str.replace(" ", "T").split(".")[0]
                event_date = datetime.fromisoformat(date_str)
            else:
                event_date = date_str
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Formato de fecha inválido: {str(e)}")

    await db.execute(text("""
        UPDATE events
        SET name = COALESCE(:name, name),
            date = COALESCE(:date, date),
            location = COALESCE(:location, location),
            format = COALESCE(:format, format),
            capacity = COALESCE(:capacity, capacity),
            price = COALESCE(:price, price),
            budget_income = COALESCE(:budget_income, budget_income),
            budget_expenses = COALESCE(:budget_expenses, budget_expenses)
        WHERE id = :id
    """), {
        "id": event_id,
        "name": data.get("name"),
        "date": event_date,
        "location": data.get("location"),
        "format": data.get("format"),
        "capacity": data.get("capacity"),
        "price": data.get("price"),
        "budget_income": data.get("budget_income"),
        "budget_expenses": data.get("budget_expenses"),
    })
    await db.commit()
    return {"ok": True}


@router.get("/psychologist/agenda")
async def get_psychologist_agenda(
    psychologist_name: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):
    """
    Obtiene la agenda personal de entrevistas y el listado de clientes asignados a una psicóloga específica.
    """
    psyc = (psychologist_name or user.get("name", "SILVI")).strip().upper()
    if "SILVI" in psyc: psyc_clean = "SILVI"
    elif "MANU" in psyc: psyc_clean = "MANU"
    elif "MAPE" in psyc: psyc_clean = "MAPE D"
    elif "ALEJA" in psyc: psyc_clean = "ALEJA"
    else: psyc_clean = psyc

    # 1. Entrevistas de agendamiento
    interviews_res = await db.execute(text("""
        SELECT ia.id, ia.user_id, ia.appointment_date, ia.time_slot, ia.status, ia.notes,
               u.name AS user_name, u.phone AS user_phone, u.client_code
        FROM interview_appointments ia
        JOIN users u ON u.id = ia.user_id
        WHERE UPPER(ia.psychologist_name) ILIKE :psyc
        ORDER BY ia.appointment_date ASC
    """), {"psyc": f"%{psyc_clean}%"})
    interviews = [{
        "id": r.id,
        "user_id": r.user_id,
        "user_name": r.user_name,
        "client_code": r.client_code or f"DL-{r.user_id:04d}",
        "phone": r.user_phone,
        "date": r.appointment_date.strftime("%d de %B, %Y") if hasattr(r.appointment_date, 'strftime') else str(r.appointment_date),
        "time": r.time_slot,
        "status": r.status,
        "notes": r.notes,
        "whatsapp_link": f"https://wa.me/{''.join(filter(str.isdigit, r.user_phone or ''))}" if r.user_phone else None
    } for r in interviews_res.fetchall()]

    # 2. Clientes asignados a esta psicóloga
    clients_res = await db.execute(text("""
        SELECT u.id, u.name, u.phone, u.client_code, p.city, p.age, p.motivacion, p.plan_tier, u.created_at
        FROM users u
        JOIN profiles p ON p.user_id = u.id
        WHERE UPPER(COALESCE(p.responsable, '')) ILIKE :psyc
        ORDER BY u.id DESC
    """), {"psyc": f"%{psyc_clean}%"})
    clients = [{
        "id": r.id,
        "name": r.name,
        "client_code": r.client_code or f"DL-{r.id:04d}",
        "phone": r.phone,
        "city": r.city or "Bogotá",
        "age": r.age,
        "plan_tier": r.plan_tier or "Sin Plan",
        "created_at": r.created_at.strftime("%d/%m/%Y") if hasattr(r.created_at, 'strftime') else (str(r.created_at)[:10] if r.created_at else "—"),
        "motivacion": r.motivacion
    } for r in clients_res.fetchall()]


    return {
        "psychologist": psyc_clean,
        "total_interviews": len(interviews),
        "total_assigned_clients": len(clients),
        "interviews": interviews,
        "assigned_clients": clients
    }



@router.get("/events/{event_id}/budget-comparison")
async def get_event_budget_comparison(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("eventos", "view"))
):
    """Compare budgeted vs actual income, expenses, and net profit for an event."""
    # Fetch event details
    ev_res = await db.execute(text("""
        SELECT name, budget_income, budget_expenses
        FROM events
        WHERE id = :id
    """), {"id": event_id})
    ev = ev_res.fetchone()
    if not ev:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
        
    # Calculate real income and expenses
    real_income = (await db.execute(text("""
        SELECT COALESCE(SUM(amount), 0) FROM income_records WHERE event_id = :id
    """), {"id": event_id})).scalar() or 0.0
    
    real_expenses = (await db.execute(text("""
        SELECT COALESCE(SUM(amount), 0) FROM expense_records WHERE event_id = :id
    """), {"id": event_id})).scalar() or 0.0
    
    b_income = float(ev.budget_income) if ev.budget_income is not None else 0.0
    b_expenses = float(ev.budget_expenses) if ev.budget_expenses is not None else 0.0
    
    b_net = b_income - b_expenses
    r_net = float(real_income) - float(real_expenses)
    
    # Calculate variance pct
    var_income = ((float(real_income) - b_income) / b_income * 100.0) if b_income > 0 else 0.0
    var_expenses = ((float(real_expenses) - b_expenses) / b_expenses * 100.0) if b_expenses > 0 else 0.0
    var_net = ((r_net - b_net) / b_net * 100.0) if b_net != 0 else 0.0
    
    verdict = "positive" if r_net >= b_net else "negative"
    
    return {
        "event_id": event_id,
        "event_name": ev.name,
        "budget": {"income": b_income, "expenses": b_expenses, "net": b_net},
        "real": {"income": float(real_income), "expenses": float(real_expenses), "net": r_net},
        "variance_pct": {"income": round(var_income, 2), "expenses": round(var_expenses, 2), "net": round(var_net, 2)},
        "verdict": verdict
    }


@router.get("/events/{event_id}/attendees")
async def get_attendees(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("eventos", "view"))
):
    """Attendees for a specific event."""
    rows = (await db.execute(text("""
        SELECT ea.user_id, ea.status, ea.created_at, u.name AS user_name, u.phone AS user_phone
        FROM event_attendees ea
        JOIN users u ON u.id = ea.user_id
        WHERE ea.event_id = :eid
        ORDER BY ea.created_at DESC
    """), {"eid": event_id})).fetchall()

    return {"attendees": [{
        "user_id": r.user_id,
        "user_name": r.user_name,
        "user_phone": r.user_phone,
        "status": r.status,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    } for r in rows]}


# ─── MATCHES ──────────────────────────────────────────────────────────────────

@router.get("/matches")
async def get_matches(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("matching", "view"))
):
    """All match requests with user and event names."""
    rows = (await db.execute(text("""
        SELECT mr.id, mr.status, mr.created_at, mr.event_id, mr.from_user, mr.to_user,
               e.name AS event_name,
               uf.name AS from_user_name,
               ut.name AS to_user_name
        FROM match_requests mr
        LEFT JOIN events e ON e.id = mr.event_id
        LEFT JOIN users uf ON uf.id = mr.from_user
        LEFT JOIN users ut ON ut.id = mr.to_user
        ORDER BY mr.created_at DESC
        LIMIT 200
    """))).fetchall()

    return {"matches": [{
        "id": r.id,
        "status": r.status,
        "event_id": r.event_id,
        "event_name": r.event_name,
        "from_user": r.from_user,
        "from_user_name": r.from_user_name,
        "to_user": r.to_user,
        "to_user_name": r.to_user_name,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    } for r in rows]}


# ─── GLOBAL SEARCH & COMPATIBILITY ───────────────────────────────────────────

@router.get("/global-search")
async def global_search(
    query: str = Query(..., min_length=1),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("dashboard", "view"))
):
    """Global search across clients, events, and employees."""
    search_pat = f"%{query}%"
    
    # 1. Search clients
    clients_res = await db.execute(text("""
        SELECT u.id, u.name, u.phone, p.motivacion
        FROM users u
        LEFT JOIN profiles p ON p.user_id = u.id
        WHERE (unaccent(u.name) ILIKE unaccent(:q) OR u.phone ILIKE :q)
          AND u.merged_into_id IS NULL
        LIMIT 10
    """), {"q": search_pat})
    clients = [{
        "id": str(r.id),
        "name": r.name,
        "phone": r.phone,
        "motivacion": r.motivacion or "Sin asignar"
    } for r in clients_res.fetchall()]
    
    # 2. Search events
    events_res = await db.execute(text("""
        SELECT id, name, date, location, format
        FROM events
        WHERE unaccent(name) ILIKE unaccent(:q) 
           OR unaccent(location) ILIKE unaccent(:q) 
           OR unaccent(format) ILIKE unaccent(:q)
        LIMIT 10
    """), {"q": search_pat})
    events = [{
        "id": r.id,
        "name": r.name,
        "date": r.date.isoformat() if r.date else None,
        "location": r.location or "Sin asignar",
        "format": r.format or "Sin asignar"
    } for r in events_res.fetchall()]
    
    # 3. Search employees
    employees_res = await db.execute(text("""
        SELECT id, full_name, role, email, phone
        FROM employees
        WHERE unaccent(full_name) ILIKE unaccent(:q) 
           OR unaccent(role) ILIKE unaccent(:q)
        LIMIT 10
    """), {"q": search_pat})
    employees = [{
        "id": str(r.id),
        "name": r.full_name,
        "role": r.role,
        "email": r.email,
        "phone": r.phone
    } for r in employees_res.fetchall()]
    
    return {
        "clients": clients,
        "events": events,
        "employees": employees
    }


@router.get("/events/{event_id}/compatibility")
async def get_event_compatibility(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("matching", "view"))
):
    """Calculates compatibility groups and matrix dynamically for an event."""
    event_res = await db.execute(text("SELECT id, name, date FROM events WHERE id = :id"), {"id": event_id})
    event = event_res.fetchone()
    if not event:
        raise HTTPException(status_code=404, detail="Evento no encontrado")
        
    attendees_res = await db.execute(text("""
        SELECT ea.user_id, u.name, p.ocean, p.apego, p.motivacion, p.rol_social, p.intereses, p.valores
        FROM event_attendees ea
        JOIN users u ON u.id = ea.user_id
        LEFT JOIN profiles p ON p.user_id = u.id
        WHERE ea.event_id = :event_id
    """), {"event_id": event_id})
    attendees = attendees_res.fetchall()
    
    if len(attendees) < 2:
        return {"groups": [], "matrix": {"names": [], "scores": []}}
        
    users_list = []
    for a in attendees:
        ocean = a.ocean if isinstance(a.ocean, dict) else (json.loads(a.ocean) if a.ocean else {})
        apego = a.apego if isinstance(a.apego, dict) else (json.loads(a.apego) if a.apego else {})
        users_list.append({
            "id": a.user_id,
            "name": a.name,
            "ocean": ocean,
            "apego": apego,
            "motivacion": a.motivacion or "exploracion",
            "rol_social": a.rol_social or "mediador",
            "intereses": a.intereses or [],
            "valores": a.valores or []
        })
        
    names = [u["name"] for u in users_list]
    n = len(users_list)
    scores_matrix = [[0] * n for _ in range(n)]
    
    for i in range(n):
        for j in range(n):
            if i == j:
                scores_matrix[i][j] = 100
                continue
            u1, u2 = users_list[i], users_list[j]
            
            ocean_score = 1.0
            if u1["ocean"] and u2["ocean"]:
                diffs = []
                for key in ["apertura", "responsabilidad", "extroversion", "amabilidad", "neuroticismo"]:
                    v1 = u1["ocean"].get(key, 0.5)
                    v2 = u2["ocean"].get(key, 0.5)
                    diffs.append(abs(v1 - v2))
                ocean_score = 1.0 - (sum(diffs) / len(diffs)) if diffs else 1.0
                
            a1 = u1["apego"].get("style", u1["apego"].get("estilo", "seguro"))
            a2 = u2["apego"].get("style", u2["apego"].get("estilo", "seguro"))
            
            attachment_compat = 0.8
            if a1 == "seguro" and a2 == "seguro":
                attachment_compat = 1.0
            elif (a1 == "ansioso" and a2 == "evitativo") or (a1 == "evitativo" and a2 == "ansioso"):
                attachment_compat = 0.35
            elif a1 == "evitativo" and a2 == "evitativo":
                attachment_compat = 0.5
            elif a1 == "ansioso" and a2 == "ansioso":
                attachment_compat = 0.6
                
            shared_interests = set(u1["intereses"]).intersection(set(u2["intereses"]))
            max_interests = max(1, len(u1["intereses"]) + len(u2["intereses"]))
            interests_score = len(shared_interests) / max_interests if shared_interests else 0.0
            
            shared_values = set(u1["valores"]).intersection(set(u2["valores"]))
            max_values = max(1, len(u1["valores"]) + len(u2["valores"]))
            values_score = len(shared_values) / max_values if shared_values else 0.0
            
            final_score = int((ocean_score * 0.4 + attachment_compat * 0.3 + interests_score * 0.15 + values_score * 0.15) * 100)
            scores_matrix[i][j] = min(100, max(0, final_score))
            
    remaining_indices = list(range(n))
    groups = []
    group_num = 1
    
    while len(remaining_indices) >= 2:
        avg_scores = []
        for idx in remaining_indices:
            other_indices = [x for x in remaining_indices if x != idx]
            if not other_indices:
                avg_scores.append((0, idx))
                continue
            avg = sum(scores_matrix[idx][o] for o in other_indices) / len(other_indices)
            avg_scores.append((avg, idx))
            
        avg_scores.sort(reverse=True)
        pivot_idx = avg_scores[0][1]
        
        remaining_indices.remove(pivot_idx)
        best_matches = []
        for idx in remaining_indices:
            score = scores_matrix[pivot_idx][idx]
            best_matches.append((score, idx))
            
        best_matches.sort(reverse=True)
        group_members_indices = [pivot_idx]
        select_count = min(3, len(remaining_indices))
        for m_idx in range(select_count):
            target_idx = best_matches[m_idx][1]
            group_members_indices.append(target_idx)
            remaining_indices.remove(target_idx)
            
        members_data = [users_list[idx] for idx in group_members_indices]
        
        all_ints = []
        for m in members_data:
            all_ints.extend(m["intereses"])
        common_ints = [item for item in set(all_ints) if all_ints.count(item) >= 2]
        
        apegos = [m["apego"].get("style", m["apego"].get("estilo", "seguro")) for m in members_data]
        seguro_count = apegos.count("seguro")
        
        exp = f"Mesa {group_num}: "
        if seguro_count >= 2:
            exp += "Alta estabilidad emocional (estilo seguro predominante). "
        else:
            exp += "Complementariedad psicológica moderada. "
            
        if common_ints:
            exp += f"Intereses en común: {', '.join(common_ints[:3])}."
        else:
            exp += "Valores y perfiles OCEAN alineados."
            
        groups.append({
            "name": f"Mesa {group_num} - Grupo IA",
            "explanation": exp,
            "members": [
                {
                    "user_id": str(m["id"]),
                    "name": m["name"],
                    "profile_summary": f"{m['apego'].get('style', m['apego'].get('estilo', 'seguro')).capitalize()} | {m['rol_social'].capitalize()}"
                } for m in members_data
            ]
        })
        group_num += 1
        
    if remaining_indices:
        idx = remaining_indices[0]
        if groups:
            groups[0]["members"].append({
                "user_id": str(users_list[idx]["id"]),
                "name": users_list[idx]["name"],
                "profile_summary": f"{users_list[idx]['apego'].get('style', users_list[idx]['apego'].get('estilo', 'seguro')).capitalize()} | {users_list[idx]['rol_social'].capitalize()}"
            })
        
    return {
        "groups": groups,
        "matrix": {
            "names": names,
            "scores": scores_matrix
        }
    }


@router.get("/events/{event_id}/export-pdf")
async def export_event_pdf(
    event_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("eventos", "view"))
):
    """Export event details (attendees, incidents, budget) as a branded Daily Lover PDF."""
    from fastapi.responses import Response
    from app.core.permissions import require_permission as _rp
    from app.services.pdf_service import build_event_pdf

    # Fetch event
    event_row = (await db.execute(text("""
        SELECT id, name, date, format, description, capacity, location, budget_income, budget_expenses
        FROM events WHERE id = :id
    """), {"id": event_id})).fetchone()

    if not event_row:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Evento no encontrado")

    event_dict = {
        "id": event_row.id,
        "name": event_row.name,
        "date": event_row.date.isoformat() if event_row.date else "",
        "format": event_row.format,
        "description": event_row.description,
        "capacity": event_row.capacity,
        "location": event_row.location,
        "budget_income": float(event_row.budget_income or 0),
        "budget_expenses": float(event_row.budget_expenses or 0),
    }

    # Fetch attendees
    att_rows = (await db.execute(text("""
        SELECT u.name, ea.status, ea.ticket_type, pef.satisfaccion
        FROM event_attendees ea
        JOIN users u ON u.id = ea.user_id
        LEFT JOIN post_event_feedback pef ON pef.event_id = ea.event_id AND pef.user_id = ea.user_id
        WHERE ea.event_id = :eid
        ORDER BY ea.status, u.name
    """), {"eid": event_id})).fetchall()

    attendees_list = [{
        "name": r.name,
        "status": r.status,
        "ticket_type": r.ticket_type,
        "satisfaccion": r.satisfaccion,
    } for r in att_rows]

    # Fetch incidents
    inc_rows = (await db.execute(text("""
        SELECT category, severity, description, resolved
        FROM event_incidents
        WHERE event_id = :eid
        ORDER BY created_at DESC
    """), {"eid": event_id})).fetchall()

    incidents_list = [{
        "category": r.category,
        "severity": r.severity,
        "description": r.description,
        "resolved": r.resolved,
    } for r in inc_rows]

    # Fetch budget comparison
    real_income = (await db.execute(text(
        "SELECT COALESCE(SUM(amount), 0) FROM income_records WHERE event_id = :eid"
    ), {"eid": event_id})).scalar() or 0.0

    real_expenses = (await db.execute(text(
        "SELECT COALESCE(SUM(amount), 0) FROM expense_records WHERE event_id = :eid"
    ), {"eid": event_id})).scalar() or 0.0

    budget_comparison = None
    if event_row.budget_income or event_row.budget_expenses:
        bi = float(event_row.budget_income or 0)
        be = float(event_row.budget_expenses or 0)
        ri = float(real_income)
        re_ = float(real_expenses)
        net_budget = bi - be
        net_real = ri - re_
        if net_real > net_budget * 1.05:
            verdict = "Por encima del presupuesto ✓"
        elif net_real < net_budget * 0.9:
            verdict = "Por debajo del presupuesto ✗"
        else:
            verdict = "Dentro del presupuesto ≈"
        budget_comparison = {
            "budget_income": bi,
            "budget_expenses": be,
            "real_income": ri,
            "real_expenses": re_,
            "verdict": verdict,
        }

    pdf_bytes = build_event_pdf(event_dict, attendees_list, incidents_list, budget_comparison)
    event_name_slug = event_dict["name"].lower().replace(" ", "_")[:20]
    filename = f"evento_{event_name_slug}_{event_id}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/historical-matches")
async def get_historical_matches(
    page: int = 1,
    limit: int = 20,
    matchmaker: str = None,
    status_filter: str = None,
    search: str = None,
    exact_name: str = None,
    user_id: Optional[int] = Query(None),
    client_code: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("matching", "view"))
):
    """Fetch historical and psychologist matches with support for exact user_id, client_code (DL-XXXX) or name matching."""
    offset = (page - 1) * limit
    where_clauses = ["1=1"]
    params = {"limit": limit, "offset": offset}

    if matchmaker and matchmaker != "all":
        where_clauses.append("matchmaker ILIKE :matchmaker")
        params["matchmaker"] = f"%{matchmaker}%"

    if status_filter and status_filter != "all":
        sf = status_filter.upper()
        if sf in ["PENDIENTE", "PENDING"]:
            where_clauses.append("(status IS NULL OR status = '' OR status ILIKE '%PENDIENTE%' OR status ~ '^[0-9.]+$' OR status ILIKE '%julio%' OR status ILIKE '%junio%' OR status ILIKE '%mayo%') AND NOT (status ILIKE '%TROUBLE%' OR status ILIKE '%REVISAR%' OR status ILIKE '%NO HAY GENTE%' OR status ILIKE '%WAITLIST%' OR status ILIKE '%REFUND%')")
        elif sf in ["APROBADO", "ACCEPTED"]:
            where_clauses.append("(status ILIKE '%APROBADO%' OR status ILIKE '%HECHO%' OR status ILIKE '%ACCEPTED%')")
        elif sf in ["RECHAZADO", "REJECTED"]:
            where_clauses.append("(status ILIKE '%RECHAZADO%' OR status ILIKE '%NOT APPROVED%' OR status ILIKE '%NO MATCH%' OR status ILIKE '%DESCALIFICADO%')")
        elif sf == "REVISAR":
            where_clauses.append("status ILIKE '%REVISAR%'")
        elif sf == "SIN_GENTE":
            where_clauses.append("(status ILIKE '%NO HAY GENTE%' OR status ILIKE '%OTRO MATCH%')")
        elif sf == "WAITLIST":
            where_clauses.append("(status ILIKE '%WAITLIST%' OR status ILIKE '%ESPERA%')")
        elif sf == "REFUND_CANCELADO":
            where_clauses.append("(status ILIKE '%REFUND%' OR status ILIKE '%CANCELADO%')")
        elif sf in ["TROUBLE", "FALLIDO"]:
            where_clauses.append("(status ILIKE '%TROUBLE%' OR status ILIKE '%REVISAR%' OR status ILIKE '%NO HAY GENTE%' OR status ILIKE '%WAITLIST%' OR status ILIKE '%REFUND%' OR status ILIKE '%CANCELADO%')")
        else:
            where_clauses.append("status ILIKE :status_filter")
            params["status_filter"] = f"%{status_filter}%"





    # 0. Auto-vinculación masiva limpia sin múltiples sentencias en un solo execute
    try:
        await db.execute(text("""
            UPDATE historical_matches hm
            SET user_id_a = u.id
            FROM users u
            WHERE hm.user_id_a IS NULL
              AND unaccent(lower(trim(hm.person_a))) = unaccent(lower(trim(u.name)))
        """))
        await db.execute(text("""
            UPDATE historical_matches hm
            SET user_id_b = u.id
            FROM users u
            WHERE hm.user_id_b IS NULL
              AND unaccent(lower(trim(hm.person_b))) = unaccent(lower(trim(u.name)))
        """))
        await db.commit()
    except Exception as e:
        logger.warning(f"Non-fatal auto-link notice: {str(e)}")


    # Filtros prioritarios por ID numérico o código DL único (ESTRICTO Y SIN MEZCLAR PERSONAS CON NOMBRES PARCIALES SIMILARES)
    if user_id:
        u_row = (await db.execute(text("SELECT id, name, client_code FROM users WHERE id = :uid"), {"uid": user_id})).fetchone()
        if u_row:
            u_name_clean = (u_row.name or "").strip()
            u_code = u_row.client_code or f"DL-{user_id:04d}"

            where_clauses.append("""(
                hm.user_id_a = :uid OR hm.user_id_b = :uid OR
                ua_id.client_code = :u_code OR ub_id.client_code = :u_code OR
                ua_name.client_code = :u_code OR ub_name.client_code = :u_code OR
                unaccent(lower(trim(hm.person_a))) = unaccent(lower(trim(:u_name))) OR
                unaccent(lower(trim(hm.person_b))) = unaccent(lower(trim(:u_name)))
            )""")
            params["uid"] = user_id
            params["u_name"] = u_name_clean
            params["u_code"] = u_code
        else:
            where_clauses.append("(hm.user_id_a = :uid OR hm.user_id_b = :uid)")
            params["uid"] = user_id

    elif client_code:
        cc_clean = client_code.strip().upper()
        u_row = (await db.execute(text("SELECT id, name FROM users WHERE upper(client_code) = :cc"), {"cc": cc_clean})).fetchone()
        u_name = (u_row.name or "").strip() if u_row else ""

        where_clauses.append("""(
            ua_id.client_code = :cc OR ub_id.client_code = :cc OR 
            ua_name.client_code = :cc OR ub_name.client_code = :cc OR
            (length(:u_name) > 2 AND (
                unaccent(lower(trim(hm.person_a))) = unaccent(lower(trim(:u_name))) OR
                unaccent(lower(trim(hm.person_b))) = unaccent(lower(trim(:u_name)))
            ))
        )""")
        params["cc"] = cc_clean
        params["u_name"] = u_name

    elif exact_name:
        en_clean = exact_name.strip()
        where_clauses.append("""(
            unaccent(lower(trim(person_a))) = unaccent(lower(trim(:exact_name))) OR
            unaccent(lower(trim(person_b))) = unaccent(lower(trim(:exact_name)))
        )""")
        params["exact_name"] = en_clean


    elif search:
        s_clean = search.strip()
        if s_clean.upper().startswith("DL-"):
            cc_clean = s_clean.upper()
            where_clauses.append("""(
                ua_id.client_code = :cc OR ub_id.client_code = :cc OR 
                ua_name.client_code = :cc OR ub_name.client_code = :cc
            )""")
            params["cc"] = cc_clean
        else:
            where_clauses.append("(unaccent(lower(person_a)) ILIKE unaccent(lower(:search)) OR unaccent(lower(person_b)) ILIKE unaccent(lower(:search)))")
            params["search"] = f"%{s_clean}%"


    where_str = " AND ".join(where_clauses)

    query_from = """
        FROM historical_matches hm
        LEFT JOIN users ua_id ON hm.user_id_a IS NOT NULL AND ua_id.id = hm.user_id_a
        LEFT JOIN profiles pa_id ON hm.user_id_a IS NOT NULL AND pa_id.user_id = hm.user_id_a
        LEFT JOIN users ua_name ON hm.user_id_a IS NULL AND unaccent(lower(trim(ua_name.name))) = unaccent(lower(trim(hm.person_a)))
        LEFT JOIN profiles pa_name ON ua_name.id IS NOT NULL AND pa_name.user_id = ua_name.id
        LEFT JOIN users ub_id ON hm.user_id_b IS NOT NULL AND ub_id.id = hm.user_id_b
        LEFT JOIN profiles pb_id ON hm.user_id_b IS NOT NULL AND pb_id.user_id = hm.user_id_b
        LEFT JOIN users ub_name ON hm.user_id_b IS NULL AND unaccent(lower(trim(ub_name.name))) = unaccent(lower(trim(hm.person_b)))
        LEFT JOIN profiles pb_name ON ub_name.id IS NOT NULL AND pb_name.user_id = ub_name.id
    """

    total_res = await db.execute(text(f"SELECT COUNT(DISTINCT hm.id) {query_from} WHERE {where_str}"), params)
    total = total_res.scalar() or 0

    rows_res = await db.execute(text(f"""
        SELECT DISTINCT
            hm.id, hm.person_a, hm.person_b, hm.matchmaker,
            hm.match_date, hm.city, hm.status, hm.observations,
            hm.user_id_a, hm.user_id_b,
            COALESCE(ua_id.client_code, ua_name.client_code) AS code_a,
            COALESCE(ub_id.client_code, ub_name.client_code) AS code_b,
            COALESCE(ua_id.name, ua_name.name, hm.person_a) AS name_a,
            COALESCE(ub_id.name, ub_name.name, hm.person_b) AS name_b,
            COALESCE(pa_id.photo_url, pa_name.photo_url) AS photo_a,
            COALESCE(pb_id.photo_url, pb_name.photo_url) AS photo_b,
            COALESCE(pa_id.city, pa_name.city, hm.city) AS city_a,
            COALESCE(pb_id.city, pb_name.city, hm.city) AS city_b
        {query_from}
        WHERE {where_str}
        ORDER BY hm.id DESC
        LIMIT :limit OFFSET :offset
    """), params)


    matches = [{
        "id": r.id,
        "person_a": r.person_a,
        "person_b": r.person_b,
        "user_id_a": r.user_id_a,
        "user_id_b": r.user_id_b,
        "code_a": r.code_a or None,
        "code_b": r.code_b or None,
        "name_a": r.name_a or r.person_a,
        "name_b": r.name_b or r.person_b,
        "photo_a": r.photo_a or None,
        "photo_b": r.photo_b or None,
        "city_a": r.city_a or r.city,
        "city_b": r.city_b or r.city,
        "matchmaker": r.matchmaker,
        "match_date": clean_excel_date_str(r.match_date) or "Por agendar",
        "city": r.city,
        "status": clean_excel_date_str(r.status) if (r.status and str(r.status).replace('.', '').isdigit()) else (r.status or "PENDIENTE"),
        "observations": r.observations,
    } for r in rows_res.fetchall()]

    target_user_diagnostic = None

    # Si hay una búsqueda de usuario (ej: "miguel" o "DL-0001") y se encuentra en la BD de usuarios registrados:
    if search and len(search.strip()) >= 2:
        search_clean = search.strip()
        usr_row = (await db.execute(text("""
            SELECT u.id, u.name, u.phone, u.client_code,
                   p.gender, p.city, p.age, p.occupation, p.photo_url, p.responsable, p.motivacion
            FROM users u
            LEFT JOIN profiles p ON p.user_id = u.id
            WHERE unaccent(lower(u.name)) ILIKE unaccent(lower(:s))
               OR unaccent(lower(COALESCE(u.client_code, ''))) ILIKE unaccent(lower(:s))
            ORDER BY u.id DESC
            LIMIT 1
        """), {"s": f"%{search_clean}%"})).fetchone()

        if usr_row:
            # 1. Evaluar completitud del formulario inicial del usuario buscado
            missing = []
            if not usr_row.name or len(usr_row.name.strip()) < 2: missing.append("Nombre")
            if not usr_row.phone: missing.append("Teléfono")
            if not usr_row.gender: missing.append("Género")
            if not usr_row.city: missing.append("Ciudad")
            if not usr_row.age: missing.append("Edad")
            if not usr_row.photo_url: missing.append("Foto de perfil")
            if not usr_row.responsable: missing.append("Psicóloga responsable")

            completeness = round(((7 - len(missing)) / 7) * 100)
            u_gender = (usr_row.gender or "").strip().lower()
            u_city = (usr_row.city or "").strip()

            target_g = "femenino" if u_gender in ["masculino", "hombre", "m"] else ("masculino" if u_gender in ["femenino", "mujer", "f"] else "ninguno")

            cands = []
            # 2. SOLO buscar candidatas/os si el usuario buscado tiene sus datos básicos y se conoce el género buscado
            if target_g != "ninguno" and u_city:
                cand_query = """
                    SELECT u.id, u.name, u.client_code, p.city, p.age, p.occupation, p.gender, p.photo_url
                    FROM users u
                    JOIN profiles p ON p.user_id = u.id
                    WHERE u.id != :target_id
                      AND u.name IS NOT NULL AND length(trim(u.name)) > 2
                      -- Exclusión estricta de cuentas de prueba
                      AND unaccent(lower(u.name)) NOT ILIKE '%test%'
                      AND unaccent(lower(u.name)) NOT ILIKE '%prueba%'
                      AND unaccent(lower(u.name)) NOT ILIKE '%consentimiento%'
                      AND unaccent(lower(u.name)) NOT ILIKE '%no terms%'
                      AND unaccent(lower(u.name)) NOT ILIKE '%demo%'
                      AND unaccent(lower(u.name)) NOT ILIKE '%dummy%'
                      AND unaccent(lower(u.name)) NOT ILIKE '%faltante%'
                      -- Requisito estricto: Perfil del candidato debe estar 100% COMPLETO (género, ciudad y edad no nulos)
                      AND p.gender IS NOT NULL AND length(trim(p.gender)) > 0
                      AND p.city IS NOT NULL AND length(trim(p.city)) > 0
                      AND p.age IS NOT NULL AND p.age > 17
                      -- Filtro estricto de Ciudad (coincidencia exacta)
                      AND unaccent(lower(trim(p.city))) = unaccent(lower(trim(:city)))
                """
                c_params = {"target_id": usr_row.id, "city": u_city}

                if target_g == "femenino":
                    cand_query += " AND lower(trim(p.gender)) IN ('femenino', 'mujer', 'f')"
                elif target_g == "masculino":
                    cand_query += " AND lower(trim(p.gender)) IN ('masculino', 'hombre', 'm')"

                cand_query += " ORDER BY u.id DESC LIMIT 6"
                cands = (await db.execute(text(cand_query), c_params)).fetchall()

            # 3. Generar propuestas algorítmicas reales SOLO SI el usuario tiene formulario completo y no tiene citas previas
            if len(matches) == 0 and len(cands) > 0:
                for idx, c in enumerate(cands, start=1):
                    matches.append({
                        "id": 90000 + c.id,
                        "person_a": usr_row.name or "Cliente",
                        "person_b": c.name,
                        "user_id_a": usr_row.id,
                        "user_id_b": c.id,
                        "code_a": usr_row.client_code or f"DL-{usr_row.id:04d}",
                        "code_b": c.client_code or f"DL-{c.id:04d}",
                        "matchmaker": f"PSICÓLOGA { (usr_row.responsable or 'SILVI').upper() }",
                        "match_date": "Propuesta Formulario Inicial",
                        "city": u_city,
                        "status": "SUGERIDO IA (FORMULARIO)",
                        "observations": f"Sugerencia estricta basada en formulario inicial completo ({u_city}, género {target_g.capitalize()})."
                    })
                total = len(matches)

            # 4. Resumen y diagnóstico estricto del formulario inicial
            if target_g == "ninguno" or not u_city or "Género" in missing or "Ciudad" in missing:
                summary_msg = f"⛔ Formulario incompleto ({completeness}%). Para proteger la calidad del proceso, este cliente NO entra al proceso de matching hasta registrar: {', '.join(missing)}."
            elif len(cands) > 0:
                summary_msg = f"✅ Formulario completo. Se identifican {len(cands)} candidatas/os reales en {u_city} que cumplen el 100% de los filtros estrictos."
            else:
                summary_msg = f"⚠️ Formulario completo ({completeness}%). No hay candidatos/as registrados en {u_city} del género {target_g.capitalize()} con perfil 100% completo en este momento."

            target_user_diagnostic = {
                "user_id": usr_row.id,
                "user_name": usr_row.name,
                "client_code": usr_row.client_code or f"DL-{usr_row.id:04d}",
                "city": u_city or "Sin registrar",
                "gender": usr_row.gender or "Sin registrar",
                "age": usr_row.age or "Sin registrar",
                "completeness": completeness,
                "missing_fields": missing,
                "compatible_candidates_count": len(cands),
                "target_gender_searched": target_g.capitalize() if target_g != "ninguno" else "Sin especificar",
                "summary": summary_msg
            }


    return {
        "matches": matches,
        "total": total,
        "page": page,
        "pages": math.ceil(total / limit) if limit > 0 else 1,
        "target_user_diagnostic": target_user_diagnostic
    }


class ReassignClientRequest(BaseModel):
    new_responsable: str
    reason: str = "Derivación clínica de caso"

@router.post("/users/{user_id}/reassign")
async def reassign_client(
    user_id: int,
    req: ReassignClientRequest,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):

    """Deriva o re-asigna un cliente a otra psicóloga o automáticamente a la siguiente disponible por menor carga."""
    user_res = await db.execute(text("SELECT id, name, phone FROM users WHERE id = :uid"), {"uid": user_id})
    u = user_res.fetchone()
    if not u:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")

    prof_res = await db.execute(text("SELECT responsable FROM profiles WHERE user_id = :uid"), {"uid": user_id})
    p = prof_res.fetchone()
    old_responsable = p.responsable if (p and p.responsable) else "Sin Asignar"

    assigned_psyc = req.new_responsable.strip()

    if assigned_psyc.upper() == "AUTO":
        psyc_list = ["Silvi", "Steffy", "Manu", "María Paula"]
        psyc_counts = []
        for psyc in psyc_list:
            cnt = (await db.execute(text("""
                SELECT COUNT(*) FROM profiles WHERE unaccent(lower(COALESCE(responsable, ''))) ILIKE unaccent(lower(:p))
            """), {"p": f"%{psyc}%"})).scalar() or 0
            psyc_counts.append((cnt, psyc))
        psyc_counts.sort()
        assigned_psyc = psyc_counts[0][1]

    await db.execute(text("""
        UPDATE profiles SET responsable = :new_psyc, updated_at = NOW() WHERE user_id = :uid
    """), {"new_psyc": assigned_psyc, "uid": user_id})

    await db.execute(text("""
        INSERT INTO reminders (title, client_name, client_phone, priority, matchmaker, due_date, notes)
        VALUES (:title, :cname, :cphone, 'ALTA', :psyc, 'Hoy', :notes)
    """), {
        "title": f"🔄 Caso Derivado de {old_responsable}: {u.name}",
        "cname": u.name,
        "cphone": u.phone or "",
        "psyc": assigned_psyc,
        "notes": f"Caso re-asignado desde {old_responsable} a {assigned_psyc}. Motivo: {req.reason}"
    })

    await db.commit()

    return {
        "ok": True,
        "user_id": user_id,
        "old_responsable": old_responsable,
        "new_responsable": assigned_psyc,
        "message": f"El cliente {u.name} ha sido re-asignado exitosamente a {assigned_psyc}."
    }


@router.get("/psychologists/performance")
async def get_psychologists_performance(
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("roles", "view"))
):
    """Módulo de Auditoría Clínico de Rendimiento para el Admin (KPIs por psicóloga)."""
    psychologists = [
        {"key": "SILVI", "name": "Silvi", "role": "Psicóloga Matchmaker Senior"},
        {"key": "STEFFY", "name": "Steffy", "role": "Psicóloga & Evaluadora Clínica"},
        {"key": "MANU", "name": "Manu", "role": "Matchmaker & Asesora de Pareja"},
        {"key": "PAULA", "name": "María Paula (MAPE)", "role": "Psicóloga & Coordinadora"}
    ]

    performance_data = []

    for p in psychologists:
        pkey = p["key"]
        pname = p["name"]

        assigned_cnt = (await db.execute(text("""
            SELECT COUNT(*) FROM profiles WHERE unaccent(lower(COALESCE(responsable, ''))) ILIKE unaccent(lower(:key))
        """), {"key": f"%{pkey}%"})).scalar() or 0

        matches_total = (await db.execute(text("""
            SELECT COUNT(*) FROM historical_matches WHERE unaccent(lower(COALESCE(matchmaker, ''))) ILIKE unaccent(lower(:key))
        """), {"key": f"%{pkey}%"})).scalar() or 0

        matches_success = (await db.execute(text("""
            SELECT COUNT(*) FROM historical_matches
            WHERE unaccent(lower(COALESCE(matchmaker, ''))) ILIKE unaccent(lower(:key))
              AND (status ILIKE '%APROBADO%' OR status ILIKE '%REALIZADA%' OR status ILIKE '%HECHO%')
        """), {"key": f"%{pkey}%"})).scalar() or 0

        notes_count = (await db.execute(text("""
            SELECT COUNT(*) FROM historical_matches
            WHERE unaccent(lower(COALESCE(matchmaker, ''))) ILIKE unaccent(lower(:key))
              AND observations IS NOT NULL AND length(trim(observations)) > 3
        """), {"key": f"%{pkey}%"})).scalar() or 0

        avg_rating_res = await db.execute(text("""
            SELECT AVG(me.chemistry_rating) FROM match_evaluations me
            JOIN historical_matches hm ON hm.id = me.match_id
            WHERE unaccent(lower(COALESCE(hm.matchmaker, ''))) ILIKE unaccent(lower(:key))
        """), {"key": f"%{pkey}%"})
        avg_rating = avg_rating_res.scalar() or 4.8

        success_rate = round((matches_success / matches_total * 100.0), 1) if matches_total > 0 else 100.0

        performance_data.append({
            "key": pkey,
            "name": pname,
            "role": p["role"],
            "assigned_clients": assigned_cnt,
            "total_matches": matches_total,
            "successful_matches": matches_success,
            "success_rate_pct": success_rate,
            "clinical_notes_logged": notes_count,
            "client_satisfaction_rating": round(float(avg_rating), 1)
        })

    return {
        "psychologists": performance_data,
        "summary": {
            "total_psychologists": len(psychologists),
            "generated_at": datetime.now().isoformat()
        }
    }


@router.patch("/historical-matches/{match_id}/status")
async def update_match_status(
    match_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("matching", "manage"))
):
    """Psychologist 1-click status update: APROBADO, PENDIENTE, RECHAZADO, TROUBLE, POSTPONED."""
    new_status = payload.get("status", "APROBADO")
    await db.execute(text("""
        UPDATE historical_matches
        SET status = :status
        WHERE id = :id
    """), {"status": new_status, "id": match_id})
    await db.commit()
    return {"message": "Estado actualizado correctamente", "id": match_id, "status": new_status}


@router.post("/historical-matches/{match_id}/send-feedback-email")
async def send_match_feedback_email(
    match_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("matching", "manage"))
):
    """Envía por correo electrónico la solicitud de evaluación post-cita obligatoria a los participantes del encuentro."""
    from app.services.email_service import send_email_html, build_feedback_email_html

    match_res = await db.execute(text("""
        SELECT hm.id, hm.person_a, hm.person_b, hm.user_id_a, hm.user_id_b,
               ua.email AS email_a, ub.email AS email_b
        FROM historical_matches hm
        LEFT JOIN users ua ON ua.id = hm.user_id_a OR unaccent(lower(trim(ua.name))) = unaccent(lower(trim(hm.person_a)))
        LEFT JOIN users ub ON ub.id = hm.user_id_b OR unaccent(lower(trim(ub.name))) = unaccent(lower(trim(hm.person_b)))
        WHERE hm.id = :mid
    """), {"mid": match_id})
    m = match_res.fetchone()
    if not m:
        raise HTTPException(status_code=404, detail="Cita no encontrada")

    sent_emails = []

    # Enviar a Persona A
    email_a = m.email_a or f"cliente_{m.user_id_a or 100}@dailylover.app"
    uid_a = m.user_id_a or 1
    html_a = build_feedback_email_html(m.person_a or "Cliente A", m.person_b or "Cliente B", m.id, uid_a)
    if send_email_html(email_a, f"🌹 Evaluación Obligatoria de Cita — {m.person_b}", html_a):
        sent_emails.append(email_a)

    # Enviar a Persona B
    email_b = m.email_b or f"cliente_{m.user_id_b or 101}@dailylover.app"
    uid_b = m.user_id_b or 2
    html_b = build_feedback_email_html(m.person_b or "Cliente B", m.person_a or "Cliente A", m.id, uid_b)
    if send_email_html(email_b, f"🌹 Evaluación Obligatoria de Cita — {m.person_a}", html_b):
        sent_emails.append(email_b)

    await db.execute(text("""
        UPDATE historical_matches
        SET feedback_email_sent_at = NOW()
        WHERE id = :mid
    """), {"mid": match_id})
    await db.commit()

    return {
        "ok": True,
        "sent_to": sent_emails,
        "message": f"Solicitud de evaluación post-cita enviada por correo electrónico a {len(sent_emails)} participante(s)."
    }



# ─── REMINDERS & PRIORITY TASKS ───────────────────────────────────────────────

@router.get("/reminders")
async def get_reminders(
    matchmaker: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("dashboard", "view"))
):
    """Retorna lista de recordatorios y tareas prioritarias asignadas a psicólogas."""
    # Ensure table exists
    await db.execute(text("""
        CREATE TABLE IF NOT EXISTS reminders (
            id SERIAL PRIMARY KEY,
            title VARCHAR(255) NOT NULL,
            client_name VARCHAR(255),
            client_phone VARCHAR(50),
            priority VARCHAR(20) DEFAULT 'ALTA', -- URGENTE, ALTA, MEDIA, BAJA
            matchmaker VARCHAR(50),              -- SILVI, STEFFY, MANU, MARÍA PAULA
            due_date VARCHAR(50),
            completed BOOLEAN DEFAULT FALSE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """))
    await db.commit()

    query = "SELECT * FROM reminders WHERE 1=1"


    params = {}
    if matchmaker:
        query += " AND (UPPER(matchmaker) = :mm OR matchmaker IS NULL)"
        params["mm"] = matchmaker.upper()

    query += " ORDER BY completed ASC, CASE priority WHEN 'URGENTE' THEN 1 WHEN 'ALTA' THEN 2 WHEN 'MEDIA' THEN 3 ELSE 4 END, id DESC"

    res = await db.execute(text(query), params)
    rows = res.fetchall()

    return {
        "reminders": [{
            "id": r.id,
            "title": r.title,
            "client_name": r.client_name,
            "client_phone": r.client_phone,
            "priority": r.priority,
            "matchmaker": r.matchmaker,
            "due_date": r.due_date,
            "completed": r.completed,
            "notes": r.notes,
            "whatsapp_link": f"https://wa.me/{''.join(filter(str.isdigit, r.client_phone or ''))}" if r.client_phone else None
        } for r in rows]
    }




@router.post("/reminders")
async def create_reminder(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("dashboard", "view"))
):
    """Crea un nuevo recordatorio o tarea prioritaria para el equipo clínico."""
    await db.execute(text("""
        INSERT INTO reminders (title, client_name, client_phone, priority, matchmaker, due_date, notes)
        VALUES (:title, :client_name, :client_phone, :priority, :matchmaker, :due_date, :notes)
    """), {
        "title": payload.get("title", "Nuevo Pendiente"),
        "client_name": payload.get("client_name"),
        "client_phone": payload.get("client_phone"),
        "priority": payload.get("priority", "ALTA"),
        "matchmaker": payload.get("matchmaker", "SILVI"),
        "due_date": payload.get("due_date", "Hoy"),
        "notes": payload.get("notes", "")
    })
    await db.commit()
    return {"message": "Recordatorio creado exitosamente"}


@router.patch("/reminders/{reminder_id}/toggle")
async def toggle_reminder(
    reminder_id: int,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("dashboard", "view"))
):
    """Marca un recordatorio como completado o pendiente."""
    await db.execute(text("""
        UPDATE reminders
        SET completed = NOT completed
        WHERE id = :id
    """), {"id": reminder_id})
    await db.commit()
    return {"message": "Estado de recordatorio actualizado"}


# ─── GESTIÓN DE DISPONIBILIDAD HORARIA POR PSICÓLOGA ───

def clean_psychologist_name(name_raw: str) -> str:
    if not name_raw: return "SILVI"
    p = name_raw.strip().upper()
    if "SILVI" in p or "SILVIA" in p: return "SILVI"
    if "STEFF" in p or "STEPH" in p: return "STEFFY"
    if "MANU" in p: return "MANU"
    if "PAULA" in p or "MAPE" in p: return "MAPE"
    if "ALEJA" in p: return "ALEJA"
    return p

@router.get("/psychologist/availability")
async def get_psychologist_availability(
    psychologist_name: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):
    """Obtiene la configuración de disponibilidad horaria por día de la semana para la psicóloga."""
    psyc_clean = clean_psychologist_name(psychologist_name or user.get("name", "SILVI"))

    res = await db.execute(text("""
        SELECT id, psychologist_name, day_of_week, start_time, end_time, slot_duration_minutes, is_active
        FROM psychologist_availability
        WHERE UPPER(psychologist_name) = :psyc
        ORDER BY day_of_week ASC
    """), {"psyc": psyc_clean})
    rows = res.fetchall()

    avail_map = {r.day_of_week: {
        "id": r.id,
        "day_of_week": r.day_of_week,
        "start_time": str(r.start_time)[:5],
        "end_time": str(r.end_time)[:5],
        "slot_duration_minutes": r.slot_duration_minutes or 45,
        "is_active": r.is_active
    } for r in rows}

    full_availability = []
    for dow in range(7):
        if dow in avail_map:
            full_availability.append(avail_map[dow])
        else:
            is_def_active = 1 <= dow <= 5
            full_availability.append({
                "id": None,
                "day_of_week": dow,
                "start_time": "09:00",
                "end_time": "17:00",
                "slot_duration_minutes": 45,
                "is_active": is_def_active
            })

    return {"psychologist": psyc_clean, "availability": full_availability}


@router.post("/psychologist/availability")
async def save_psychologist_availability(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: dict = Depends(require_permission("clientes", "view"))
):
    """Guarda o actualiza la franja horaria y estado (activo/inactivo) de un día para la psicóloga."""
    psyc_clean = clean_psychologist_name(payload.get("psychologist_name") or user.get("name", "SILVI"))

    dow = int(payload.get("day_of_week", 1))
    is_active = bool(payload.get("is_active", True))
    start_str = payload.get("start_time", "09:00")
    end_str = payload.get("end_time", "17:00")
    duration = int(payload.get("slot_duration_minutes", 45))

    t_start = datetime.strptime(start_str, "%H:%M").time()
    t_end = datetime.strptime(end_str, "%H:%M").time()

    exist = (await db.execute(text("""
        SELECT id FROM psychologist_availability
        WHERE UPPER(psychologist_name) = :psyc AND day_of_week = :dow
    """), {"psyc": psyc_clean, "dow": dow})).fetchone()

    if exist:
        await db.execute(text("""
            UPDATE psychologist_availability
            SET start_time = :s, end_time = :e, slot_duration_minutes = :dur, is_active = :act
            WHERE id = :id
        """), {"s": t_start, "e": t_end, "dur": duration, "act": is_active, "id": exist.id})
    else:
        await db.execute(text("""
            INSERT INTO psychologist_availability (psychologist_name, day_of_week, start_time, end_time, slot_duration_minutes, is_active)
            VALUES (:psyc, :dow, :s, :e, :dur, :act)
        """), {"psyc": psyc_clean, "dow": dow, "s": t_start, "e": t_end, "dur": duration, "act": is_active})

    await db.commit()

    return {"ok": True, "message": f"Disponibilidad del día {dow} guardada con éxito."}


CITY_NORM_MAP = {
    "bgota": "Bogotá",
    "bogota": "Bogotá",
    "medellin": "Medellín",
    "baq": "Barranquilla",
    "bquilla": "Barranquilla",
    "quilla": "Barranquilla",
    "barranca": "Barranquilla",
    "bmanga": "Bucaramanga",
    "buc": "Bucaramanga",
    "buca": "Bucaramanga",
    "bga": "Bucaramanga",
    "ctg": "Cartagena",
    "mad": "Madrid",
    "mia": "Miami",
    "cdmx": "CDMX",
    "peira": "Pereira",
    "ibag": "Ibagué",
    "caqu": "Caquetá",
}

def normalize_city_name(raw_city: Optional[str]) -> str:
    if not raw_city:
        return ""
    c = str(raw_city).strip()
    if not c or c in (",", "2 Dates", "Todo El Mundo"):
        return ""
    c_lower = c.lower()
    if c_lower in CITY_NORM_MAP:
        return CITY_NORM_MAP[c_lower]
    return c.title()

def normalize_pref(raw_orientation: Optional[str], raw_gender: Optional[str]) -> str:
    ori = str(raw_orientation or "").lower().strip()
    gen = str(raw_gender or "").lower().strip()

    if "bi" in ori:
        return "bi"
    if "lesb" in ori:
        return "lesb"
    if "gay" in ori or "homo" in ori:
        if "fem" in gen or "mujer" in gen or "female" in gen:
            return "lesb"
        return "gay"
    if "straight" in ori or "hetero" in ori:
        return "hetero"
    
    return ""


@router.get("/clients-plans")
async def get_clients_plans(
    request: Request,
    x_api_key: Optional[str] = Header(None, alias="x-api-key"),
    db: AsyncSession = Depends(get_db)
):
    """
    Endpoint de solo lectura para consultar la lista de clientes con su plan_tier, ciudad y preferencia.
    Diseñado para integrarse periódicamente con Google Sheets (Apps Script).
    Protegido vía API Key (header x-api-key o SHEET_INTEGRATION_API_KEY).
    """
    api_key_env = (os.environ.get("SHEET_INTEGRATION_API_KEY") or get_settings().sheet_integration_api_key or "").strip()
    provided_key = (
        x_api_key or 
        request.headers.get("x-api-key") or 
        request.headers.get("X-API-Key") or 
        request.headers.get("X-Api-Key") or ""
    ).strip()

    if not api_key_env or not provided_key or provided_key != api_key_env:
        raise HTTPException(status_code=401, detail="API key inválida")

    result = await db.execute(text("""
        SELECT u.name, u.email, u.phone, p.plan_tier, p.city, p.orientation, p.gender
        FROM profiles p
        JOIN users u ON u.id = p.user_id
        WHERE p.plan_tier IS NOT NULL 
          AND TRIM(p.plan_tier) != ''
        ORDER BY u.name ASC
    """))
    rows = result.fetchall()

    clients = [
        {
            "name": row.name or "",
            "email": row.email or "",
            "phone": row.phone or "",
            "plan_tier": row.plan_tier,
            "city": normalize_city_name(row.city),
            "pref": normalize_pref(row.orientation, row.gender)
        }
        for row in rows
    ]

    return {
        "clients": clients,
        "updated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    }




